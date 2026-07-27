#!/usr/bin/env python3
###########################################################################
#
# Copyright 2026 Samsung Electronics All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
###########################################################################

"""Tests for heap_analyzer.py. Run: python3 -m unittest test_heap_analyzer -v"""

import importlib.util
import io
import json
import os
import tempfile
import unittest
from contextlib import redirect_stdout, redirect_stderr
from types import SimpleNamespace
from unittest.mock import patch

MODULE_PATH = os.path.join(os.path.dirname(__file__), "heap_analyzer.py")
SPEC = importlib.util.spec_from_file_location("heap_analyzer", MODULE_PATH)
heap_analyzer = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(heap_analyzer)


class TempDirTestCase(unittest.TestCase):
    def setUp(self):
        self._tempdir = tempfile.TemporaryDirectory()
        self.workdir = self._tempdir.name
        self.addCleanup(self._tempdir.cleanup)

    def make_log(self, content, name="input.log"):
        path = os.path.join(self.workdir, name)
        with open(path, "w", encoding="utf-8") as logfile:
            logfile.write(content)
        return path

    def output_dir(self, name="output"):
        return os.path.join(self.workdir, name)

    def run_main(self, *argv):
        stdout, stderr = io.StringIO(), io.StringIO()
        with patch.object(heap_analyzer.os, "getcwd", return_value=self.workdir), \
                redirect_stdout(stdout), redirect_stderr(stderr):
            code = heap_analyzer.main(list(argv))
        return code, stdout.getvalue(), stderr.getvalue()


# ========================================================================
# [S2] OutputManager
# ========================================================================

class OutputManagerTest(TempDirTestCase):
    def test_prepare_creates_output_and_extracted_dirs(self):
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()

        self.assertTrue(os.path.isdir(output.root))
        self.assertTrue(os.path.isdir(output.extracted_dir))
        self.assertEqual(os.listdir(output.extracted_dir), [])

    def test_prepare_wipes_previous_content(self):
        root = self.output_dir()
        stale_file = os.path.join(root, "old_report.txt")
        stale_subdir = os.path.join(root, "extracted", "stale")
        os.makedirs(stale_subdir)
        with open(stale_file, "w") as stale:
            stale.write("stale")

        heap_analyzer.OutputManager(root).prepare()

        self.assertFalse(os.path.exists(stale_file))
        self.assertFalse(os.path.exists(stale_subdir))
        self.assertTrue(os.path.isdir(os.path.join(root, "extracted")))

    def test_prepare_refuses_unsafe_locations(self):
        for unsafe in (os.sep, os.path.expanduser("~"), os.getcwd()):
            with self.subTest(path=unsafe):
                with self.assertRaises(ValueError):
                    heap_analyzer.OutputManager(unsafe).prepare()

    def test_prepare_refuses_existing_non_directory(self):
        not_a_dir = os.path.join(self.workdir, "occupied")
        with open(not_a_dir, "w") as occupied:
            occupied.write("file, not dir")

        with self.assertRaises(ValueError):
            heap_analyzer.OutputManager(not_a_dir).prepare()

    def test_path_helpers(self):
        output = heap_analyzer.OutputManager(self.output_dir())

        self.assertEqual(
            output.path("report.txt"), os.path.join(output.root, "report.txt")
        )
        self.assertEqual(
            output.extracted_event_path(3, "alloc_fail"),
            os.path.join(output.extracted_dir, "event_03_alloc_fail.log"),
        )


# ========================================================================
# [S3] Log Extractor
# ========================================================================

TS = "[2026-07-22 16:10:46.014] "
TS2 = "[2026-07-22 16:10:56.599] "

FAIL_EVENT = f"""{TS}mm_malloc: Allocation failed!!! We dont have enough memory. Try to free dead task stack areas
{TS}mm_ioctl_garbagecollection: Fail to call sched_garbagecollection, errno 38
{TS}mm_manage_alloc_fail_dump: Allocation failed from user heap.
{TS}mm_manage_alloc_fail_dump:  - requested size 137088
{TS2}REGION #0 Start=0x0x60287080, End=0x0x6077fff0, Size=5214080
"""

HEAPINFO_EVENT = f"""{TS}TASH>>heapinfo
{TS}
{TS}****************************************************************
{TS}     KERNEL HEAP INFORMATION
{TS}****************************************************************
{TS}Total                           : 60135168 (100%)
"""


class ExtractorTest(TempDirTestCase):
    def extract(self, content):
        log = self.make_log(content)
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()
        return heap_analyzer.extract_events(log, output)

    def read_extracted(self, event):
        with open(event.extracted_path, encoding="utf-8") as extracted:
            return extracted.read()

    def test_alloc_fail_event_with_context_lines(self):
        events = self.extract("boot noise\n" + FAIL_EVENT + "trailing noise\n")

        self.assertEqual(len(events), 1)
        event = events[0]
        self.assertEqual(event.kind, "alloc_fail")
        self.assertEqual(event.source_line_start, 2)  # mm_malloc context line
        self.assertEqual(event.end_reason, "eof")
        self.assertEqual(event.time_start, "2026-07-22 16:10:46.014")
        self.assertEqual(event.time_end, "2026-07-22 16:10:56.599")

        content = self.read_extracted(event)
        self.assertNotIn("[2026-07-22", content)          # timestamps stripped
        self.assertIn("mm_malloc: Allocation failed", content)
        self.assertIn("trailing noise", content)
        self.assertTrue(content.startswith("mm_malloc:"))

    def test_glued_start_marker_is_split_mid_line(self):
        # Pattern (1): event start glued to the tail of the previous line.
        log = (
            "mm_malloc: Allocation failed!!! Try to free dead tas"
            "mm_manage_alloc_fail_dump: Allocation failed from user heap.\n"
            "mm_manage_alloc_fail_dump:  - requested size 70135168\n"
        )
        events = self.extract(log)

        self.assertEqual(len(events), 1)
        lines = self.read_extracted(events[0]).splitlines()
        self.assertEqual(lines[0], "mm_malloc: Allocation failed!!! Try to free dead tas")
        self.assertTrue(lines[1].startswith("mm_manage_alloc_fail_dump: Allocation failed"))

    def test_heapinfo_event_records_command(self):
        events = self.extract(HEAPINFO_EVENT)

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0].kind, "heapinfo")
        self.assertEqual(events[0].command, "heapinfo")
        content = self.read_extracted(events[0])
        self.assertIn("KERNEL HEAP INFORMATION", content)
        self.assertIn("Total", content)

    def test_multiple_events_are_split_into_files(self):
        second = HEAPINFO_EVENT.replace("TASH>>heapinfo", "TASH>>heapinfo -a")
        events = self.extract(HEAPINFO_EVENT + second + FAIL_EVENT)

        self.assertEqual(len(events), 3)
        self.assertEqual([event.kind for event in events],
                         ["heapinfo", "heapinfo", "alloc_fail"])
        self.assertEqual(events[0].end_reason, "next_event")
        self.assertEqual(events[1].command, "heapinfo -a")
        self.assertEqual(
            [os.path.basename(event.extracted_path) for event in events],
            ["event_01_heapinfo.log", "event_02_heapinfo.log",
             "event_03_alloc_fail.log"],
        )

    def test_kernel_banner_without_tash_echo_splits_heapinfo_runs(self):
        banner_only = "".join(
            line + "\n" for line in HEAPINFO_EVENT.splitlines()[1:]
        )
        events = self.extract(HEAPINFO_EVENT + banner_only)

        self.assertEqual(len(events), 2)
        self.assertEqual(events[0].command, "heapinfo")
        self.assertIsNone(events[1].command)
        self.assertIn("KERNEL HEAP INFORMATION", self.read_extracted(events[1]))

    def test_reboot_terminates_event(self):
        # Pattern (4): dump cut short by a reboot.
        log = FAIL_EVENT + (
            f"{TS2}boardctl: Board will Reboot now. pid: 65\n"
            f"{TS2}ROM:[V1.1]\n"
            f"{TS2}FLASHRATE:1\n"
        )
        events = self.extract(log)

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0].end_reason, "reboot")
        content = self.read_extracted(events[0])
        self.assertNotIn("boardctl", content)
        self.assertNotIn("ROM:[V1.1]", content)

    def test_tash_prompt_terminates_event(self):
        events = self.extract(HEAPINFO_EVENT + "TASH>>hello\nHello, World!!\n")

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0].end_reason, "tash_prompt")
        self.assertNotIn("Hello, World!!", self.read_extracted(events[0]))

    def test_log_without_timestamps(self):
        stripped = "".join(
            heap_analyzer.TIMESTAMP_RE.sub("", line) + "\n"
            for line in FAIL_EVENT.splitlines()
        )
        events = self.extract(stripped)

        self.assertEqual(len(events), 1)
        self.assertIsNone(events[0].time_start)
        self.assertIn("requested size 137088", self.read_extracted(events[0]))

    def test_no_events_in_plain_log(self):
        self.assertEqual(self.extract("TASH>>hello\njust some noise\n"), [])


class RealLogExtractorTest(TempDirTestCase):
    """End-to-end extraction against the real captured logs, if present."""

    def extract_file(self, name):
        path = os.path.join(os.path.dirname(MODULE_PATH), name)
        if not os.path.isfile(path):
            self.skipTest(f"{name} not available")
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()
        return heap_analyzer.extract_events(path, output)

    def test_alloc_fail_test_log(self):
        events = self.extract_file("alloc_fail_test.log")

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0].kind, "alloc_fail")
        self.assertEqual(events[0].end_reason, "eof")

    def test_alloc_fail_real_log(self):
        events = self.extract_file("alloc_fail_real.log")

        self.assertEqual(len(events), 1)
        event = events[0]
        self.assertEqual(event.kind, "alloc_fail")
        self.assertEqual(event.end_reason, "reboot")
        self.assertEqual(event.source_line_start, 3)   # mm_malloc context
        self.assertEqual(event.source_line_end, 5170)  # cut by reboot at 5171
        self.assertEqual(event.time_start, "2026-07-22 16:10:46.014")

    def test_heapinfo_test_log(self):
        events = self.extract_file("heapinfo_test.log")

        self.assertEqual(len(events), 2)
        self.assertEqual([event.kind for event in events],
                         ["heapinfo", "heapinfo"])
        self.assertEqual(events[0].command, "heapinfo")
        self.assertEqual(events[1].command, "heapinfo -a")
        self.assertEqual(events[0].end_reason, "next_event")
        with open(events[1].extracted_path, encoding="utf-8") as extracted:
            content = extracted.read()
        self.assertIn("REGION #0", content)
        self.assertIn("< by Alive Threads >", content)


# ========================================================================
# [S4] Section Parsers
# ========================================================================

FAIL_DUMP = """mm_malloc: Allocation failed!!! We dont have enough memory.
mm_manage_alloc_fail_dump: Allocation failed from user heap.
mm_manage_alloc_fail_dump:  - requested size 137088
mm_manage_alloc_fail_dump:  - caller address = 0x0e5721bd
mm_manage_alloc_fail_dump:  - largest free size : 112064
mm_manage_alloc_fail_dump:  - total free size   : 625632
mm_manage_alloc_fail_dump: *************************************************
mm_manage_alloc_fail_dump:              Summary of current app (app1) heap memory usage
mm_manage_alloc_fail_dump: *************************************************

mm_check_heap_corruption: Heap start = 0x60287080 end = 0x6077fff0
mm_check_heap_corruption: No heap corruption detected
****************************************************************
REGION #0 Start=0x0x60287080, End=0x0x6077fff0, Size=5214080
****************************************************************
  MemAddr |   Size   | Status |    Owner   |  Pid  |
----------|----------|--------|------------|-------|
0x60287080 |       16 |   A    | 0x    dead |  15   |
0x60287090 |     8224 |   A    | 0x e00c809 |  16(S)|
0x602890b0 |      992 |   F    |            |       |
** PID(S) in Pid column means that mem is used for stack of PID

****************************************************************
     Summary of Heap Usages (Size in Bytes)
****************************************************************
Total                           : 5214080 (100%)
  - Allocated (Current / Peak)  : 705936 (62%) / 1031088 (91%)
  - Free (Current)              : 425712 (37%)
  - Reserved                    : 32
"""


class ParserTestBase(TempDirTestCase):
    def parse(self, content, kind="alloc_fail", command=None):
        path = self.make_log(content, name="event.log")
        extracted = heap_analyzer.ExtractedEvent(
            index=1, kind=kind, command=command, extracted_path=path
        )
        return heap_analyzer.parse_event(extracted)


class SectionParserTest(ParserTestBase):
    def test_fail_info_fields(self):
        event = self.parse(FAIL_DUMP)
        info = event.fail_info

        self.assertEqual(info.heap_kind, "user")
        self.assertEqual(info.requested_size, 137088)
        self.assertEqual(info.caller_address, "0x0e5721bd")
        self.assertEqual(info.largest_free_size, 112064)
        self.assertEqual(info.total_free_size, 625632)

    def test_heap_snapshot_from_app_banner(self):
        event = self.parse(FAIL_DUMP)

        self.assertEqual([heap.name for heap in event.heaps], ["app1"])
        heap = event.heaps[0]
        self.assertEqual(heap.heap_range, ("0x60287080", "0x6077fff0"))
        self.assertIs(heap.corruption, False)
        self.assertEqual(len(heap.regions), 1)
        self.assertEqual(heap.regions[0].start, "0x60287080")
        self.assertEqual(heap.regions[0].size, 5214080)

    def test_dump_table_blocks(self):
        heap = self.parse(FAIL_DUMP).heaps[0]

        self.assertEqual(len(heap.blocks), 3)
        self.assertFalse(heap.blocks_truncated)
        first, stack, free = heap.blocks
        self.assertEqual((first.address, first.size, first.status),
                         ("0x60287080", 16, "A"))
        self.assertEqual(first.owner, "0xdead")
        self.assertEqual(first.pid, "15")
        self.assertEqual(stack.category, "stack")
        self.assertTrue(stack.is_stack)
        self.assertEqual(free.category, "free")
        self.assertIsNone(free.owner)

    def test_heap_summary_fields(self):
        summary = self.parse(FAIL_DUMP).heaps[0].summary

        self.assertEqual(summary.total, 5214080)
        self.assertEqual(summary.alloc_current, 705936)
        self.assertEqual(summary.alloc_peak, 1031088)
        self.assertEqual(summary.free_current, 425712)
        self.assertEqual(summary.reserved, 32)

    def test_truncated_dump_row_sets_flag(self):
        # Pattern (2): last row cut mid-field, then the next section starts.
        truncated = FAIL_DUMP.replace(
            "0x602890b0 |      992 |   F    |            |       |\n"
            "** PID(S) in Pid column means that mem is used for stack of PID\n",
            "0x602890b0 |     \n",
        )
        event = self.parse(truncated)
        heap = event.heaps[0]

        self.assertEqual(len(heap.blocks), 2)
        self.assertTrue(heap.blocks_truncated)
        self.assertEqual(event.diagnostics.malformed_rows, 1)
        self.assertIn("dump_table:app1", event.diagnostics.truncated_sections)

    def test_partial_dump_row_followed_by_noise_sets_flag(self):
        truncated = FAIL_DUMP.replace(
            "0x602890b0 |      992 |   F    |            |       |\n"
            "** PID(S) in Pid column means that mem is used for stack of PID\n",
            "0x602890b0 |     \n"
            "interleaved noise\n",
        )
        event = self.parse(truncated)
        heap = event.heaps[0]

        self.assertTrue(heap.blocks_truncated)
        self.assertIn("dump_table:app1", event.diagnostics.truncated_sections)

    def test_dump_rows_outside_region_are_skipped(self):
        for bad_address in ("0x60287070", "0x6077fff0"):
            with self.subTest(address=bad_address):
                content = FAIL_DUMP.replace(
                    "0x60287080 |       16 |   A    | 0x    dead |  15   |",
                    f"{bad_address} |       32 |   A    | 0x    dead |  15   |",
                )
                event = self.parse(content)
                heap = event.heaps[0]

                self.assertEqual(len(heap.blocks), 2)
                self.assertEqual(event.diagnostics.malformed_rows, 1)

    def test_table_cut_at_eof_sets_flag(self):
        cut = FAIL_DUMP.split("** PID(S)")[0]
        heap = self.parse(cut).heaps[0]

        self.assertEqual(len(heap.blocks), 3)
        self.assertTrue(heap.blocks_truncated)

    def test_malformed_row_mid_table_is_not_truncation(self):
        malformed = FAIL_DUMP.replace(
            "0x60287090 |     8224 |   A    | 0x e00c809 |  16(S)|",
            "0x60287090 |     8224 |   A    |            |  16   |",  # A without owner
        )
        event = self.parse(malformed)
        heap = event.heaps[0]

        self.assertEqual(len(heap.blocks), 2)
        self.assertFalse(heap.blocks_truncated)
        self.assertEqual(event.diagnostics.malformed_rows, 1)

    def test_kernel_banner_switches_heap(self):
        content = FAIL_DUMP + """mm_manage_alloc_fail_dump: *****************
mm_manage_alloc_fail_dump:            Summary of Kernel heap memory usage
mm_manage_alloc_fail_dump: *****************

mm_check_heap_corruption: Heap start = 0x60126900 end = 0x63a7fff0
mm_check_heap_corruption: No heap corruption detected

****************************************************************
     Summary of Heap Usages (Size in Bytes)
****************************************************************
Total                           : 60135168 (100%)
  - Allocated (Current / Peak)  : 752464 (1%) / 9131136 (15%)
  - Free (Current)              : 59382704 (98%)
  - Reserved                    : 32
"""
        event = self.parse(content)

        self.assertEqual([heap.name for heap in event.heaps], ["app1", "kernel"])
        kernel = event.heaps[1]
        self.assertEqual(kernel.heap_range, ("0x60126900", "0x63a7fff0"))
        self.assertEqual(kernel.summary.total, 60135168)
        self.assertEqual(event.heaps[0].summary.total, 5214080)

    def test_heapinfo_banner_switches_heap(self):
        content = """TASH>>heapinfo

****************************************************************
     KERNEL HEAP INFORMATION
****************************************************************
mm_check_heap_corruption: Heap start = 0x60126900 end = 0x63a7fff0
mm_check_heap_corruption: No heap corruption detected

****************************************************************
     Summary of Heap Usages (Size in Bytes)
****************************************************************
Total                           : 60135168 (100%)
  - Allocated (Current / Peak)  : 750512 (1%) / 9131296 (15%)
  - Free (Current)              : 59384656 (98%)
  - Reserved                    : 32

****************************************************************
     app1 HEAP INFORMATION
****************************************************************
mm_check_heap_corruption: Heap start = 0x63a812d0 end = 0x63f7fff0
mm_check_heap_corruption: No heap corruption detected

****************************************************************
     Summary of Heap Usages (Size in Bytes)
****************************************************************
Total                           : 5238064 (100%)
  - Allocated (Current / Peak)  : 67056 (1%) / 67056 (1%)
  - Free (Current)              : 5171008 (98%)
  - Reserved                    : 32
"""
        event = self.parse(content, kind="heapinfo", command="heapinfo")

        self.assertEqual([heap.name for heap in event.heaps], ["kernel", "app1"])
        self.assertEqual(event.heaps[0].summary.total, 60135168)
        self.assertEqual(event.heaps[1].summary.total, 5238064)
        self.assertEqual(event.heaps[1].heap_range, ("0x63a812d0", "0x63f7fff0"))


DETAILS_SECTION = """****************************************************************
     Details of Heap Usages (Size in Bytes)
****************************************************************
< Free >
  - Number of Free Node               : 27
  - Largest Free Node Size            : 421008

< Allocation >
  - Current Size (Alive Allocation) = (1) + (2) + (3)
     . by Dead Threads (*) (1)        : 178592
     . by Alive Threads
        - Sum of "STACK"(**) (2)      : 68096
        - Sum of "CURR_HEAP" (3)      : 459216
** NOTE **
(*)  Alive allocation by dead threads might be used by others or might be a leakage.
(**) Only Idle task has a separate stack region,
  rest are all allocated on the heap region.

Available fragmented memory segments in heap memory
Nodelist[0] ranging [1, 32] : num 0, size 0 [Bytes]
Nodelist[1] ranging [33, 64] : num 1, size 32 [Bytes]
Nodelist[18] ranging [4194305, 8388608] : num 1, size 5161040 [Bytes]

< by Dead Threads >
 Pid | Size
-----|------
  18 |  1824

< by Alive Threads >
PID |  PPID | STACK | CURR_HEAP | PEAK_HEAP |    BIN   | NAME
----|-------|-------|-----------|-----------|----------|----------
  0 |     0 |  1024 |        -1 |        -1 |   kernel | CPU0 IDLE()
 19 |    18 |  8168 |     16592 |     25152 |     app1 | app1()
 27 |    19 |  4064 |         0 |         0 |     app1 | wifi msg handler()
"""

ASSERT_TAIL = """security level: 0
===========================================================
Assertion details
===========================================================
Assertion failed CPU0 at file: mm_heap/mm_manage_allocfail.c line 209 task: hello pid: 29
print_assert_detail: Assert location (PC) : 0x0e02bfe7
check_assert_location: Code asserted in normal thread!
===========================================================
Asserted task's stack details
===========================================================
check_sp_corruption: Current SP is User Thread SP: 63a93e48
check_sp_corruption: User stack:
print_stack_dump:   base: 63a93fa0
print_stack_dump:   size: 00001ff0
print_stack_dump:   used: 0000095c
up_stackdump: 63a93e40: xxxxxxxx xxxxxxxx 000000d1 60105508 601167fc 0e02bcfb 0e0aa80d 004fed30
up_stackdump: 63a93e60: 00000001 004ec050 00012ce0 000000df 00000013 0000005f 63a80020 00000000
===========================================================
Asserted task's TCB info
===========================================================
task_show_tcbinfo: State       : 4
task_show_tcbinfo: Flags       : 1056
task_show_tcbinfo: Syscall 0   : 0xe30d2e2
===========================================================
List of all tasks in the system
===========================================================
                           NAME |   PID |  PRI |    USED / TOTAL STACK | STACK ALLOC ADDR | TCB ADDR | TASK STATE
-------------------------------------------------------------------------------------------------------------------
                      CPU0 IDLE |     0 |    0 |     736 /    1024 |       0x60126500 | 0x6010c014 |          3
                          hello |    29 |  100 |    2396 /    8176 |       0x63a91fa0 | 0x601ddc30 |          4
-------------------------------------------------------------------------------------------------------------------
===========================================================
Loading location information
===========================================================
elf_show_all_bin_section_addr: [common] Text Addr : 0xe261010, Text Size : 8679424
elf_show_all_bin_section_addr: [app1] Text Addr : 0xeaa8030, Text Size : 4894720
===========================================================
Checking app 1 heap for corruption
===========================================================
mm_holder: -1
mm_check_heap_corruption: Heap start = 0x63a812d0 end = 0x63f7fff0
mm_check_heap_corruption: No heap corruption detected
"""


class DetailSectionParserTest(ParserTestBase):
    def test_details_fields(self):
        details = self.parse(FAIL_DUMP + DETAILS_SECTION).heaps[0].details

        self.assertEqual(details.free_node_count, 27)
        self.assertEqual(details.largest_free_node, 421008)
        self.assertEqual(details.dead_thread_alloc, 178592)
        self.assertEqual(details.stack_sum, 68096)
        self.assertEqual(details.curr_heap_sum, 459216)

    def test_nodelist_buckets(self):
        nodelist = self.parse(FAIL_DUMP + DETAILS_SECTION).heaps[0].nodelist

        self.assertEqual(len(nodelist), 3)
        self.assertEqual((nodelist[1].range_min, nodelist[1].range_max,
                          nodelist[1].num, nodelist[1].size), (33, 64, 1, 32))
        self.assertEqual(nodelist[2].size, 5161040)

    def test_dead_and_alive_threads(self):
        heap = self.parse(FAIL_DUMP + DETAILS_SECTION).heaps[0]

        self.assertEqual(heap.dead_threads, [(18, 1824)])
        self.assertEqual(len(heap.alive_threads), 3)
        idle, app1, wifi = heap.alive_threads
        self.assertEqual((idle.pid, idle.name, idle.bin), (0, "CPU0 IDLE", "kernel"))
        self.assertEqual((app1.curr_heap, app1.peak_heap), (16592, 25152))
        self.assertEqual(wifi.name, "wifi msg handler")

    def test_assertion_info(self):
        event = self.parse(FAIL_DUMP + ASSERT_TAIL)
        info = event.assertion

        self.assertEqual(info.file, "mm_heap/mm_manage_allocfail.c")
        self.assertEqual(info.line, 209)
        self.assertEqual(info.task, "hello")
        self.assertEqual(info.pid, 29)
        self.assertEqual(info.pc, "0x0e02bfe7")
        self.assertEqual(info.tcb["State"], "4")
        self.assertEqual(info.tcb["Syscall 0"], "0xe30d2e2")

    def test_stack_info(self):
        stack = self.parse(FAIL_DUMP + ASSERT_TAIL).assertion.stack

        self.assertEqual(stack.sp, "63a93e48")
        self.assertEqual(stack.base, "63a93fa0")
        self.assertEqual(stack.size, 0x1ff0)
        self.assertEqual(stack.used, 0x95c)
        self.assertFalse(stack.truncated)

    def test_glued_stackdump_marks_truncation(self):
        # Pattern (3): stack dump cut mid-line and glued to the next header.
        glued = ASSERT_TAIL.replace(
            "up_stackdump: 63a93e60: 00000001 004ec050 00012ce0 000000df "
            "00000013 0000005f 63a80020 00000000",
            "up_stackdump: 63a93e60: 00000001 004ec050 604===========",
        )
        event = self.parse(FAIL_DUMP + glued)

        self.assertTrue(event.assertion.stack.truncated)
        self.assertIn("stack_dump", event.diagnostics.truncated_sections)

    def test_task_list(self):
        task_list = self.parse(FAIL_DUMP + ASSERT_TAIL).task_list

        self.assertFalse(task_list.truncated)
        self.assertEqual(len(task_list.entries), 2)
        hello = task_list.entries[1]
        self.assertEqual((hello.name, hello.pid, hello.priority), ("hello", 29, 100))
        self.assertEqual((hello.stack_used, hello.stack_total), (2396, 8176))
        self.assertEqual(hello.state, 4)

    def test_task_list_truncated_row(self):
        # Pattern (4): task row cut short right before the reboot.
        cut = ASSERT_TAIL.replace(
            "                          hello |    29 |  100 |    2396 /    8176 "
            "|       0x63a91fa0 | 0x601ddc30 |          4",
            "    WIFI_WATCHDOG_PERIODIC_TASK |    37 |  100 |    1036 /    2048 |    ",
        )
        event = self.parse(FAIL_DUMP + cut)

        self.assertTrue(event.task_list.truncated)
        self.assertEqual(len(event.task_list.entries), 1)
        self.assertIn("task_list", event.diagnostics.truncated_sections)

    def test_loading_info(self):
        loading = self.parse(FAIL_DUMP + ASSERT_TAIL).loading_info

        self.assertEqual(len(loading), 2)
        self.assertEqual(loading[0].name, "common")
        self.assertEqual(loading[0].text_addr, "0xe261010")
        self.assertEqual(loading[1].text_size, 4894720)

    def test_assert_tail_corruption_check_does_not_clobber_heap(self):
        event = self.parse(FAIL_DUMP + ASSERT_TAIL)
        # The re-check at the very end reports app heap 0x63a812d0, which
        # must not overwrite the app1 snapshot range from the dump section.
        self.assertEqual(event.heaps[0].heap_range, ("0x60287080", "0x6077fff0"))


class RealLogParserTest(TempDirTestCase):
    """Parse the real captured logs end-to-end (extract + parse)."""

    def parse_file(self, name):
        path = os.path.join(os.path.dirname(MODULE_PATH), name)
        if not os.path.isfile(path):
            self.skipTest(f"{name} not available")
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()
        extracted = heap_analyzer.extract_events(path, output)
        return [heap_analyzer.parse_event(item) for item in extracted]

    def test_alloc_fail_test_log_full_parse(self):
        event = self.parse_file("alloc_fail_test.log")[0]

        self.assertEqual(event.fail_info.requested_size, 70135168)
        self.assertEqual(event.fail_info.caller_address, "0x0e273d2f")
        self.assertEqual(event.fail_info.total_free_size, 5161040)

        names = [heap.name for heap in event.heaps]
        self.assertEqual(names, ["app1", "kernel"])
        app1 = event.heaps[0]
        self.assertFalse(app1.blocks_truncated)
        free_blocks = [b for b in app1.blocks if b.status == "F"]
        self.assertEqual(len(free_blocks), 1)
        self.assertEqual(free_blocks[0].size, 5161040)
        self.assertEqual(len(app1.blocks), 104)
        # Observed table sum (77008) differs from the reported
        # alloc_current (77024) by heap metadata; M5 cross-checks the two.
        self.assertEqual(sum(b.size for b in app1.blocks if b.status == "A"),
                         77008)
        self.assertEqual(app1.summary.total, 5238064)
        self.assertEqual(event.heaps[1].summary.total, 60135168)

        # M4 sections
        self.assertEqual(app1.details.dead_thread_alloc, 1824)
        self.assertEqual(app1.details.stack_sum, 55472)
        self.assertEqual(app1.dead_threads, [(18, 1824)])
        self.assertEqual(len(app1.nodelist), 19)
        self.assertEqual(app1.nodelist[18].size, 5161040)
        self.assertEqual(event.heaps[1].details.free_node_count, 2)

        assertion = event.assertion
        self.assertEqual(assertion.task, "hello")
        self.assertEqual(assertion.pid, 29)
        self.assertEqual(assertion.pc, "0x0e02bfe7")
        self.assertEqual(assertion.stack.used, 0x95c)
        self.assertFalse(assertion.stack.truncated)
        self.assertEqual(len(event.task_list.entries), 26)
        self.assertFalse(event.task_list.truncated)
        self.assertEqual([b.name for b in event.loading_info],
                         ["common", "app1"])

    def test_alloc_fail_real_log_detects_truncation(self):
        event = self.parse_file("alloc_fail_real.log")[0]

        self.assertEqual(event.fail_info.requested_size, 137088)
        app1 = event.heaps[0]
        self.assertEqual(app1.name, "app1")
        self.assertTrue(app1.blocks_truncated)
        self.assertGreaterEqual(event.diagnostics.malformed_rows, 1)
        self.assertGreater(len(app1.blocks), 1000)
        self.assertIn("dump_table:app1", event.diagnostics.truncated_sections)

        # M4 sections: assert tail with truncated stack dump and task list
        assertion = event.assertion
        self.assertEqual(assertion.task, "SpeechDetectorWorker")
        self.assertEqual(assertion.pid, 65)
        self.assertTrue(assertion.stack.truncated)
        self.assertTrue(event.task_list.truncated)
        self.assertGreater(len(event.task_list.entries), 30)
        self.assertIn("stack_dump", event.diagnostics.truncated_sections)
        self.assertIn("task_list", event.diagnostics.truncated_sections)

    def test_heapinfo_log_full_parse(self):
        events = self.parse_file("heapinfo_test.log")

        plain, full = events
        self.assertEqual([heap.name for heap in plain.heaps], ["kernel", "app1"])
        self.assertEqual(plain.heaps[0].summary.total, 60135168)
        self.assertEqual(plain.heaps[0].blocks, [])
        self.assertEqual(len(plain.heaps[1].alive_threads), 25)
        self.assertEqual(plain.heaps[1].alive_threads[-1].name,
                         "ble msg handler")

        kernel, app1 = full.heaps
        self.assertGreater(len(kernel.blocks), 100)
        self.assertGreater(len(app1.blocks), 50)
        self.assertFalse(app1.blocks_truncated)
        free = [b for b in app1.blocks if b.status == "F"]
        self.assertEqual(len(free), 1)
        self.assertEqual(free[0].size, 5171008)


# ========================================================================
# [S1] Models
# ========================================================================

class ModelsTest(unittest.TestCase):
    def test_diagnostics_mark_truncated_deduplicates(self):
        diagnostics = heap_analyzer.Diagnostics()
        diagnostics.mark_truncated("dump_table")
        diagnostics.mark_truncated("dump_table")
        diagnostics.mark_truncated("task_list")

        self.assertEqual(diagnostics.truncated_sections, ["dump_table", "task_list"])

    def test_analysis_result_serializes_to_json(self):
        event = heap_analyzer.Event(index=1, kind=heap_analyzer.EVENT_ALLOC_FAIL)
        event.heaps.append(heap_analyzer.HeapSnapshot(name="app1"))
        result = heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=10, total_events=1),
            events=[event],
        )

        payload = json.loads(
            json.dumps(
                heap_analyzer.asdict(result), default=heap_analyzer._json_default
            )
        )
        self.assertEqual(payload["events"][0]["kind"], "alloc_fail")
        self.assertEqual(payload["events"][0]["heaps"][0]["name"], "app1")


# ========================================================================
# [S5] Analyzer
# ========================================================================

class AnalyzerTest(ParserTestBase):
    def analyze(self, content):
        event = self.parse(content)
        heap_analyzer.analyze_event(event)
        return event

    def test_by_pid_and_owner_aggregation(self):
        derived = self.analyze(FAIL_DUMP).heaps[0].derived

        self.assertEqual(derived.total_allocated, 16 + 8224)
        self.assertEqual(derived.observed_free, 992)
        self.assertEqual(derived.by_pid["15"]["total_allocated"], 16)
        self.assertEqual(derived.by_pid["16"]["stack_size"], 8224)
        self.assertEqual(derived.by_pid["16"]["heap_size"], 0)
        self.assertEqual(derived.by_owner["0xdead"]["pids"], ["15"])
        self.assertEqual(derived.largest_allocation["size"], 8224)

    def test_histogram_buckets(self):
        histogram = self.analyze(FAIL_DUMP).heaps[0].derived.histogram

        self.assertEqual(histogram[0]["label"], "1-16")
        self.assertEqual(histogram[0]["count"], 1)
        self.assertEqual(histogram[-1]["label"], "8193-16384")
        self.assertEqual(histogram[-1]["stack_count"], 1)

    def test_fragmentation_from_fail_info(self):
        derived = self.analyze(FAIL_DUMP).heaps[0].derived
        # No details section: falls back to the fail header numbers.
        self.assertAlmostEqual(derived.fragmentation_index,
                               (1 - 112064 / 625632) * 100, places=2)

    def test_fragmentation_prefers_details(self):
        derived = self.analyze(FAIL_DUMP + DETAILS_SECTION).heaps[0].derived
        self.assertAlmostEqual(derived.fragmentation_index,
                               (1 - 421008 / 425712) * 100, places=2)

    def test_partial_flag_from_truncated_dump(self):
        truncated = FAIL_DUMP.replace(
            "0x602890b0 |      992 |   F    |            |       |\n"
            "** PID(S) in Pid column means that mem is used for stack of PID\n",
            "0x602890b0 |     \n",
        )
        derived = self.analyze(truncated).heaps[0].derived

        self.assertTrue(derived.partial)

    def test_pid_task_name_join(self):
        event = heap_analyzer.Event(index=1, kind="alloc_fail")
        heap = heap_analyzer.HeapSnapshot(name="app1")
        heap.blocks.append(heap_analyzer.Block(
            region=0, address="0x63a81000", size=64, status="A",
            owner="0xe001", pid="19",
        ))
        heap.alive_threads.append(heap_analyzer.ThreadEntry(
            pid=19, ppid=18, stack=8168, curr_heap=16592,
            peak_heap=25152, bin="app1", name="app1",
        ))
        event.heaps.append(heap)

        heap_analyzer.analyze_event(event)

        self.assertEqual(heap.derived.by_pid["19"]["name"], "app1")
        self.assertEqual(heap.derived.by_pid["19"]["bin"], "app1")
        self.assertEqual(heap.derived.largest_allocation["task"], "app1")

    def test_continuity_gap_detection(self):
        heap = heap_analyzer.HeapSnapshot(name="x")
        for address, size in (("0x00001000", 16), ("0x00001100", 16),
                              ("0x00001108", 16)):
            heap.blocks.append(heap_analyzer.Block(
                region=0, address=address, size=size, status="A",
                owner="0xe001", pid="1",
            ))

        gaps = heap_analyzer._find_continuity_gaps(heap)

        self.assertEqual([gap.kind for gap in gaps], ["gap", "overlap"])
        self.assertEqual(gaps[0].start, "0x1010")
        self.assertEqual(gaps[0].end, "0x1100")

    def test_heapinfo_without_dump_has_no_block_metrics(self):
        content = ("TASH>>heapinfo\n"
                   "     Summary of Heap Usages (Size in Bytes)\n"
                   "Total                           : 1000000 (100%)\n"
                   "  - Free (Current)              : 425712 (42%)\n"
                   + DETAILS_SECTION)
        event = self.parse(content, kind="heapinfo", command="heapinfo")
        heap_analyzer.analyze_event(event)
        derived = event.heaps[0].derived

        self.assertIsNone(derived.total_allocated)
        self.assertEqual(derived.histogram, [])
        self.assertAlmostEqual(derived.fragmentation_index,
                               (1 - 421008 / 425712) * 100, places=2)


# ========================================================================
# [S7] Reporters (M1: console/json skeleton)
# ========================================================================

class ReporterTest(TempDirTestCase):
    def test_console_report_no_events(self):
        result = heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=123, total_events=0)
        )
        text = heap_analyzer.render_console_report(result)

        self.assertIn("Events found: 0", text)
        self.assertIn("No alloc-fail or heapinfo events", text)

    def test_console_report_lists_events_with_truncation(self):
        event = heap_analyzer.Event(
            index=1, kind="alloc_fail", time_start="2026-07-22 16:10:46.014",
            source_line_start=3, source_line_end=5170,
        )
        event.diagnostics.mark_truncated("dump_table")
        result = heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=1, total_events=1),
            events=[event],
        )
        text = heap_analyzer.render_console_report(result)

        self.assertIn("--- EVENT LIST ---", text)
        self.assertIn("alloc_fail", text)
        self.assertIn("3-5170", text)
        self.assertIn("dump_table", text)

    def test_format_size(self):
        self.assertEqual(heap_analyzer.format_size(None), "N/A")
        self.assertEqual(heap_analyzer.format_size(512), "512.00 B")
        self.assertEqual(heap_analyzer.format_size(5238064), "5.00 MB")

    def build_wide_event(self):
        """One owner spread over 8 PIDs, 7 owners total, a long task name."""
        event = heap_analyzer.Event(index=1, kind="alloc_fail")
        heap = heap_analyzer.HeapSnapshot(name="app1")
        address = 0x60287080
        for pid in (3, 21, 47, 5, 116, 65, 48, 9):     # shared owner
            heap.blocks.append(heap_analyzer.Block(
                region=0, address=f"0x{address:08x}", size=32, status="A",
                owner="0xe5721bd", pid=str(pid)))
            address += 32
        for extra in range(6):                          # 6 more owners
            heap.blocks.append(heap_analyzer.Block(
                region=0, address=f"0x{address:08x}", size=16, status="A",
                owner=f"0xe00c80{extra}", pid="1"))
            address += 16
        heap.alive_threads.append(heap_analyzer.ThreadEntry(
            pid=116, ppid=1, stack=2048, curr_heap=0, peak_heap=0,
            bin="kernel", name="WIFI_WATCHDOG_PERIODIC_TASK"))
        event.heaps.append(heap)
        heap_analyzer.analyze_event(event)
        return event

    def render(self, event):
        result = heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=1,
                                            total_events=1),
            events=[event],
        )
        return heap_analyzer.render_console_report(result)

    def test_owner_pid_list_is_never_elided(self):
        report = self.render(self.build_wide_event())

        self.assertNotIn("...", report)
        owner_line = next(line for line in report.splitlines()
                          if line.strip().startswith("0xe5721bd"))
        # All 8 PIDs, numerically sorted.
        self.assertIn("3, 5, 9, 21, 47, 48, 65, 116", owner_line)

    def test_all_owners_are_listed(self):
        report = self.render(self.build_wide_event())

        for extra in range(6):
            self.assertIn(f"0xe00c80{extra}", report)

    def test_long_task_name_is_not_truncated(self):
        report = self.render(self.build_wide_event())

        self.assertIn("WIFI_WATCHDOG_PERIODIC_TASK [kernel]", report)

    def test_owner_line_shows_raw_addr2line_result(self):
        event = self.build_wide_event()
        resolved = {"function": "osif_timer_wrapper", "resolved": True,
                    "location": "/root/tizenrt/os/kernel/osif_timer.c:120"}
        for block in event.heaps[0].blocks:
            if block.owner == "0xe5721bd":
                block.symbol = resolved

        report = self.render(event)

        owner_line = next(line for line in report.splitlines()
                          if line.strip().startswith("0xe5721bd"))
        # addr2line output verbatim: function + full location.
        self.assertIn(
            "<osif_timer_wrapper /root/tizenrt/os/kernel/osif_timer.c:120>",
            owner_line)

    def test_owner_line_shows_unresolved_addr2line_result_verbatim(self):
        # A binary existed but addr2line could not resolve this address:
        # the raw "?? ??:0" answer is shown, not hidden.
        event = self.build_wide_event()
        for block in event.heaps[0].blocks:
            if block.owner == "0xe5721bd":
                block.symbol = {"function": "??", "resolved": False,
                                "location": "??:0"}

        report = self.render(event)

        owner_line = next(line for line in report.splitlines()
                          if line.strip().startswith("0xe5721bd"))
        self.assertIn("<?? ??:0>", owner_line)

    def test_owner_line_without_binary_shows_owner_only(self):
        # Symbols never resolved (no binary): no code annotation at all.
        report = self.render(self.build_wide_event())

        owner_line = next(line for line in report.splitlines()
                          if line.strip().startswith("0xe5721bd"))
        self.assertNotIn("<", owner_line)
        self.assertNotIn("??", owner_line)


# ========================================================================
# [S6] Symbol Resolver
# ========================================================================

class SymbolResolverTest(ParserTestBase):
    def build_result(self):
        event = self.parse(FAIL_DUMP + ASSERT_TAIL)
        heap_analyzer.analyze_event(event)
        return heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=1,
                                            total_events=1),
            events=[event],
        )

    def make_bin_dir(self, *names):
        bin_dir = os.path.join(self.workdir, "bin")
        os.makedirs(bin_dir, exist_ok=True)
        for name in names:
            with open(os.path.join(bin_dir, name), "w") as elf:
                elf.write("fake elf")
        return bin_dir

    def test_no_elf_files_skips_resolution(self):
        result = self.build_result()
        empty = os.path.join(self.workdir, "empty")
        os.makedirs(empty)
        with patch.object(heap_analyzer.os.path, "isfile", return_value=False):
            heap_analyzer.resolve_symbols(result, empty, "addr2line")

        self.assertFalse(result.symbols_resolved)
        self.assertIn("no ELF binaries", result.symbols_skip_reason)

    def test_batch_resolution_attaches_symbols(self):
        result = self.build_result()
        bin_dir = self.make_bin_dir("tinyara.axf")
        elf = os.path.join(bin_dir, "tinyara.axf")
        addresses = heap_analyzer._collect_addresses(result)
        stdout = "".join(f"func_{i}\nsource_{i}.c:{i}\n"
                         for i in range(len(addresses)))

        with patch.object(heap_analyzer, "_find_elf_files",
                          return_value=[elf]), \
             patch.object(heap_analyzer.subprocess, "run",
                          return_value=SimpleNamespace(stdout=stdout)) as run:
            heap_analyzer.resolve_symbols(result, bin_dir, "addr2line")

        self.assertEqual(run.call_count, 1)          # one batch per ELF
        self.assertTrue(result.symbols_resolved)
        event = result.events[0]
        self.assertTrue(event.fail_info.caller_symbol["resolved"])
        self.assertTrue(event.assertion.pc_symbol["resolved"])
        block = event.heaps[0].blocks[0]
        self.assertTrue(block.symbol["resolved"])
        self.assertEqual(block.symbol["elf"], "tinyara.axf")

    def test_resolved_answer_wins_over_unknown(self):
        result = self.build_result()
        bin_dir = self.make_bin_dir("tinyara.axf", "common_dbg")
        elves = [os.path.join(bin_dir, "tinyara.axf"),
                 os.path.join(bin_dir, "common_dbg")]
        addresses = heap_analyzer._collect_addresses(result)
        unknown = "??\n??:0\n" * len(addresses)
        resolved = "".join(f"known_{i}\nknown_{i}.c:{i}\n"
                           for i in range(len(addresses)))
        outputs = [SimpleNamespace(stdout=unknown),
                   SimpleNamespace(stdout=resolved)]

        with patch.object(heap_analyzer, "_find_elf_files",
                          return_value=elves), \
             patch.object(heap_analyzer.subprocess, "run",
                          side_effect=outputs):
            heap_analyzer.resolve_symbols(result, bin_dir, "addr2line")

        block = result.events[0].heaps[0].blocks[0]
        self.assertTrue(block.symbol["resolved"])
        self.assertEqual(block.symbol["elf"], "common_dbg")

    def test_no_symbols_flag(self):
        result = self.build_result()
        heap_analyzer.resolve_symbols(result, "anywhere", "addr2line",
                                      no_symbols=True)

        self.assertEqual(result.symbols_skip_reason,
                         "disabled by --no-symbols")


# ========================================================================
# [S7] Excel / HTML reporters
# ========================================================================

class ExcelHtmlReporterTest(ParserTestBase):
    def build_result(self):
        event = self.parse(FAIL_DUMP + DETAILS_SECTION + ASSERT_TAIL)
        heap_analyzer.analyze_event(event)
        return heap_analyzer.AnalysisResult(
            source=heap_analyzer.SourceInfo(path="x.log", size=1,
                                            total_events=1),
            events=[event],
        )

    def test_excel_sheets(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        result = self.build_result()
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()

        heap_analyzer.report_excel(result, output)

        workbook = load_workbook(output.path("report.xlsx"))
        names = workbook.sheetnames
        self.assertIn("Events", names)
        self.assertIn("Summary", names)
        self.assertIn("app1_PID", names)
        self.assertIn("app1_Blocks", names)
        self.assertIn("Tasks", names)
        self.assertIn("Diagnostics", names)

        events = list(workbook["Events"].values)
        self.assertEqual(events[1][1], "alloc_fail")
        blocks = list(workbook["app1_Blocks"].values)
        self.assertEqual(blocks[1][1], "0x60287080")

    def test_html_map_content(self):
        result = self.build_result()
        output = heap_analyzer.OutputManager(self.output_dir()).prepare()

        heap_analyzer.report_html(result, output)

        with open(output.path("memory_map.html"), encoding="utf-8") as html:
            content = html.read()
        self.assertIn("TizenRT Heap Memory Map", content)
        self.assertIn('"address": "0x60287080"', content)
        self.assertIn('"category": "unknown"', content)  # inferred tail
        self.assertIn('"fragmentation_index"', content)
        self.assertIn('id="strip"', content)             # vertical strip
        self.assertIn("rowHeight", content)
        self.assertNotIn("showHeap", content)            # no checkboxes
        self.assertNotIn("__DATA__", content)
        # middle detail panel: per-run member blocks + click rendering
        self.assertIn('id="detail"', content)
        self.assertIn("renderDetail", content)
        self.assertIn('"blocks":', content)
        self.assertIn('"location":', content)   # addr2line info per block
        self.assertIn("codeOf", content)        # raw addr2line rendering
        # right-hand report panel with divider, fed by the embedded report
        self.assertIn('class="divider"', content)
        self.assertIn('id="reportPre"', content)
        self.assertIn('"report":', content)
        self.assertIn("[FAIL] allocation from user heap", content)

    def test_html_runs_merge_contiguous_same_category(self):
        heap = heap_analyzer.HeapSnapshot(name="x")
        heap.regions.append(heap_analyzer.Region(
            number=0, start="0x00001000", end="0x00001100", size=0x100))
        rows = (("0x00001000", 16, "A", False), ("0x00001010", 32, "A", False),
                ("0x00001030", 16, "A", True), ("0x00001040", 32, "F", False),
                ("0x00001060", 16, "A", False))
        for address, size, status, is_stack in rows:
            heap.blocks.append(heap_analyzer.Block(
                region=0, address=address, size=size, status=status,
                owner="0xe001" if status == "A" else None,
                pid="1" if status == "A" else None, is_stack=is_stack,
                category="stack" if is_stack else (
                    "heap" if status == "A" else "free"),
            ))

        runs = heap_analyzer._html_runs(heap)

        self.assertEqual(
            [(r["category"], r["size"], r["count"]) for r in runs],
            [("heap", 48, 2),        # two contiguous heap allocs merged
             ("stack", 16, 1),
             ("free", 32, 1),
             ("heap", 16, 1),
             ("unknown", 0x100 - 0x70, 1)],  # inferred tail
        )
        self.assertEqual(runs[0]["address"], "0x00001000")
        # Merged runs keep their member allocations for the detail panel.
        self.assertEqual([b["address"] for b in runs[0]["blocks"]],
                         ["0x00001000", "0x00001010"])
        self.assertEqual(runs[0]["blocks"][0]["pid"], "1")
        self.assertEqual(runs[0]["blocks"][0]["owner"], "0xe001")
        self.assertEqual(runs[2]["blocks"], [])   # free: no allocations

    def test_html_gap_inference(self):
        result = self.build_result()
        heap = result.events[0].heaps[0]
        directory = heap_analyzer._pid_directory(result.events[0])

        blocks = heap_analyzer._display_blocks(heap, directory)

        # 16 + 8224 + 992 parsed bytes, then one inferred gap to region end.
        gaps = [b for b in blocks if b["category"] == "gap"]
        self.assertEqual(len(gaps), 1)
        self.assertEqual(gaps[0]["address"], "0x60289490")
        region = heap.regions[0]
        self.assertEqual(gaps[0]["size"],
                         int(region.end, 16) - 0x60289490)


class RealLogReportTest(TempDirTestCase):
    """Full pipeline over real logs, checking the rendered report."""

    def report_for(self, name):
        path = os.path.join(os.path.dirname(MODULE_PATH), name)
        if not os.path.isfile(path):
            self.skipTest(f"{name} not available")
        code, stdout, stderr = self.run_main(path, "--no-symbols")
        self.assertEqual(code, 0, stderr)
        return stdout

    def test_alloc_fail_report_content(self):
        report = self.report_for("alloc_fail_test.log")

        self.assertIn("[FAIL] allocation from user heap", report)
        self.assertIn("request exceeds total free memory", report)
        self.assertIn("-- HEAP [app1]", report)
        self.assertIn("-- HEAP [kernel]", report)
        self.assertIn("RecorderWorker", report)          # PID-task join
        self.assertIn("Largest allocation", report)
        self.assertIn("[ASSERT] mm_heap/mm_manage_allocfail.c:209", report)
        self.assertIn("Dead-thread allocations", report)

    def test_truncated_report_warns(self):
        report = self.report_for("alloc_fail_real.log")

        self.assertIn("(!) TRUNCATED sections:", report)
        self.assertIn("dump_table:app1", report)
        self.assertIn("LOWER BOUND", report)
        self.assertIn("no free node is large enough", report)
        # No elided owner/PID lists anywhere in the report.
        self.assertNotIn(", ...", report)
        owner_line = next(line for line in report.splitlines()
                          if line.strip().startswith("0xe5721bd"))
        pid_count = len(owner_line.split("  ")[-1].split(","))
        self.assertGreater(pid_count, 6)   # previously capped at 6 + "..."

    def test_heapinfo_report_content(self):
        report = self.report_for("heapinfo_test.log")

        self.assertIn("EVENT #1: heapinfo  (command: heapinfo)", report)
        self.assertIn("EVENT #2: heapinfo  (command: heapinfo -a)", report)
        self.assertIn("-- HEAP [kernel]", report)
        self.assertNotIn("[FAIL]", report)


# ========================================================================
# [S8] CLI & main (end-to-end skeleton)
# ========================================================================

class CliTest(TempDirTestCase):
    def test_missing_logfile_returns_error(self):
        code, _, stderr = self.run_main(
            os.path.join(self.workdir, "no_such.log"),
        )

        self.assertEqual(code, 1)
        self.assertIn("not found", stderr)

    def test_output_option_is_rejected(self):
        with redirect_stderr(io.StringIO()), self.assertRaises(SystemExit) as raised:
            heap_analyzer.build_arg_parser().parse_args(
                ["input.log", "--output", "other"]
            )

        self.assertEqual(raised.exception.code, 2)

    def test_smoke_run_creates_artifacts(self):
        log = self.make_log("TASH>>hello\nno heap events here\n")
        out = self.output_dir()
        code, stdout, _ = self.run_main(log, "--no-symbols")

        self.assertEqual(code, 0)
        self.assertIn("Events found: 0", stdout)
        self.assertTrue(os.path.isfile(os.path.join(out, "report.txt")))
        self.assertTrue(os.path.isfile(os.path.join(out, "analysis.json")))
        self.assertTrue(os.path.isdir(os.path.join(out, "extracted")))

        with open(os.path.join(out, "analysis.json"), encoding="utf-8") as dump:
            payload = json.load(dump)
        self.assertEqual(payload["source"]["total_events"], 0)
        self.assertEqual(payload["events"], [])

    def test_rerun_wipes_previous_artifacts(self):
        log = self.make_log("nothing\n")
        out = self.output_dir()
        self.run_main(log, "--no-symbols")
        leftover = os.path.join(out, "extracted", "event_99_stale.log")
        with open(leftover, "w") as stale:
            stale.write("stale")

        code, _, _ = self.run_main(log, "--no-symbols")

        self.assertEqual(code, 0)
        self.assertFalse(os.path.exists(leftover))

    def test_event_option_out_of_range(self):
        log = self.make_log("nothing\n")
        code, _, stderr = self.run_main(
            log, "--event", "1", "--no-symbols"
        )

        self.assertEqual(code, 1)
        self.assertIn("out of range", stderr)


if __name__ == "__main__":
    unittest.main()
