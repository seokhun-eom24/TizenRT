#!/usr/bin/env python3
###########################################################################
#
# Copyright 2026 Samsung Electronics All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
###########################################################################

"""Analyze TizenRT heap logs (allocation-failure dumps and heapinfo output).

Pipeline:
    [S0] constants/regex -> [S1] models -> [S2] output dir ->
    [S3] extractor -> [S4] section parsers -> [S5] analyzer ->
    [S6] symbols -> [S7] reporters -> [S8] CLI

Every artifact (extracted per-event logs, report.txt, report.xlsx,
memory_map.html, analysis.json) is written under the output directory,
which is wiped and recreated on every run.
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from collections import deque
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Dict, List, Optional, Tuple


# ========================================================================
# [S0] Constants & Regex
# ========================================================================

EVENT_ALLOC_FAIL = "alloc_fail"
EVENT_HEAPINFO = "heapinfo"

# "[2026-07-22 16:10:46.014] " prefix; logs may also come without it.
TIMESTAMP_RE = re.compile(r"^\[(\d{4}-\d{2}-\d{2} [0-9:.]+)\]\s?")

# --- Event boundaries (Stage 1) -----------------------------------------
# Alloc-fail events start at the mm_manage_alloc_fail marker, which can be
# glued to the tail of the previous line (serial logs mix streams).
FAIL_START_RE = re.compile(
    r"(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):"
    r"\s*Allocation failed from (user|kernel) heap\."
)
# Lines that immediately precede the marker and belong to the same failure.
FAIL_CONTEXT_RE = re.compile(
    r"mm_malloc:.*Allocation failed|mm_ioctl_garbagecollection:"
)
# heapinfo events start at the TASH command echo (command recorded) or, if
# the echo is missing, at the KERNEL HEAP INFORMATION banner.
TASH_HEAPINFO_RE = re.compile(r"TASH>>\s*(heapinfo\b.*?)\s*$")
KERNEL_BANNER_RE = re.compile(r"KERNEL HEAP INFORMATION")
# A TASH prompt inside an event means the shell is back: the event is over.
TASH_PROMPT_RE = re.compile(r"TASH>>")
# The board rebooting also terminates whatever event was being dumped.
REBOOT_RE = re.compile(
    r"boardctl: Board will Reboot|^ROM:\[V|Autoboot in \d+ milliseconds"
)

DEFAULT_OUTPUT_DIR = "output"
EXTRACTED_DIR_NAME = "extracted"
REPORT_TXT_NAME = "report.txt"
REPORT_XLSX_NAME = "report.xlsx"
HTML_MAP_NAME = "memory_map.html"
ANALYSIS_JSON_NAME = "analysis.json"


# ========================================================================
# [S1] Data Models
# ========================================================================

class ParseStatus(str, Enum):
    OK = "ok"
    TRUNCATED = "truncated"
    MALFORMED_PARTIAL = "malformed_partial"


@dataclass
class Region:
    number: int
    start: str
    end: str
    size: int


@dataclass
class Block:
    region: int
    address: str
    size: int
    status: str                      # "A" | "F" | "G"(inferred gap)
    owner: Optional[str] = None
    pid: Optional[str] = None
    is_stack: bool = False
    category: str = "heap"           # heap | stack | free | gap
    inferred: bool = False
    symbol: Optional[dict] = None    # filled by symbol resolver


@dataclass
class HeapSummary:
    total: Optional[int] = None
    alloc_current: Optional[int] = None
    alloc_peak: Optional[int] = None
    free_current: Optional[int] = None
    reserved: Optional[int] = None


@dataclass
class HeapDetails:
    free_node_count: Optional[int] = None
    largest_free_node: Optional[int] = None
    dead_thread_alloc: Optional[int] = None
    stack_sum: Optional[int] = None
    curr_heap_sum: Optional[int] = None


@dataclass
class NodeBucket:
    range_min: int
    range_max: int
    num: int
    size: int


@dataclass
class ThreadEntry:
    pid: int
    ppid: int
    stack: int
    curr_heap: int
    peak_heap: int
    bin: str
    name: str


@dataclass
class TaskEntry:
    name: str
    pid: int
    priority: int
    stack_used: int
    stack_total: int
    stack_addr: str
    tcb_addr: str
    state: int


@dataclass
class TaskList:
    entries: List[TaskEntry] = field(default_factory=list)
    truncated: bool = False


@dataclass
class StackInfo:
    sp: Optional[str] = None
    base: Optional[str] = None
    size: Optional[int] = None
    used: Optional[int] = None
    truncated: bool = False


@dataclass
class AssertionInfo:
    file: Optional[str] = None
    line: Optional[int] = None
    task: Optional[str] = None
    pid: Optional[int] = None
    pc: Optional[str] = None
    pc_symbol: Optional[dict] = None
    stack: Optional[StackInfo] = None
    tcb: Dict[str, str] = field(default_factory=dict)


@dataclass
class BinTextAddr:
    name: str                        # "common", "app1", ...
    text_addr: str
    text_size: int


@dataclass
class FailInfo:
    heap_kind: Optional[str] = None  # "user" | "kernel"
    requested_size: Optional[int] = None
    caller_address: Optional[str] = None
    caller_symbol: Optional[dict] = None
    largest_free_size: Optional[int] = None
    total_free_size: Optional[int] = None


@dataclass
class ContinuityGap:
    region: int
    start: str
    end: str
    kind: str                        # "gap" | "overlap"


@dataclass
class Diagnostics:
    truncated_sections: List[str] = field(default_factory=list)
    malformed_rows: int = 0
    unknown_lines: int = 0

    def mark_truncated(self, section: str):
        if section not in self.truncated_sections:
            self.truncated_sections.append(section)


@dataclass
class DerivedMetrics:
    by_pid: Dict[str, dict] = field(default_factory=dict)
    by_owner: Dict[str, dict] = field(default_factory=dict)
    top_pids: List[dict] = field(default_factory=list)
    top_owners: List[dict] = field(default_factory=list)
    histogram: List[dict] = field(default_factory=list)
    fragmentation_index: Optional[float] = None
    total_allocated: Optional[int] = None
    observed_free: Optional[int] = None
    largest_allocation: Optional[dict] = None
    continuity_gaps: List[ContinuityGap] = field(default_factory=list)
    partial: bool = False            # metrics computed from truncated data


@dataclass
class HeapSnapshot:
    name: str                        # "kernel" | "app1" | ...
    heap_range: Optional[Tuple[str, str]] = None
    corruption: Optional[bool] = None
    regions: List[Region] = field(default_factory=list)
    blocks: List[Block] = field(default_factory=list)
    blocks_truncated: bool = False
    summary: Optional[HeapSummary] = None
    details: Optional[HeapDetails] = None
    nodelist: List[NodeBucket] = field(default_factory=list)
    dead_threads: List[Tuple[int, int]] = field(default_factory=list)
    alive_threads: List[ThreadEntry] = field(default_factory=list)
    derived: Optional[DerivedMetrics] = None


@dataclass
class ExtractedEvent:
    """Stage 1 output: one cleaned event log saved under output/extracted/."""
    index: int
    kind: str                        # EVENT_ALLOC_FAIL | EVENT_HEAPINFO
    command: Optional[str] = None    # e.g. "heapinfo -a"
    time_start: Optional[str] = None
    time_end: Optional[str] = None
    source_line_start: int = 0
    source_line_end: int = 0
    extracted_path: Optional[str] = None
    end_reason: str = "eof"          # eof | next_event | reboot | tash_prompt


@dataclass
class Event:
    index: int
    kind: str
    command: Optional[str] = None
    time_start: Optional[str] = None
    time_end: Optional[str] = None
    source_line_start: int = 0
    source_line_end: int = 0
    extracted_path: Optional[str] = None
    end_reason: str = "eof"
    fail_info: Optional[FailInfo] = None
    heaps: List[HeapSnapshot] = field(default_factory=list)
    assertion: Optional[AssertionInfo] = None
    task_list: Optional[TaskList] = None
    loading_info: List[BinTextAddr] = field(default_factory=list)
    diagnostics: Diagnostics = field(default_factory=Diagnostics)


@dataclass
class SourceInfo:
    path: str
    size: int
    total_events: int = 0


@dataclass
class AnalysisResult:
    source: SourceInfo
    events: List[Event] = field(default_factory=list)
    symbols_resolved: bool = False
    symbols_skip_reason: Optional[str] = None


# ========================================================================
# [S2] Output Manager
# ========================================================================

class OutputManager:
    """Owns the output directory: wipes it on prepare(), hands out paths."""

    def __init__(self, root):
        self.root = os.path.abspath(root)

    def prepare(self):
        self._check_safe_to_wipe()
        if os.path.isdir(self.root):
            shutil.rmtree(self.root)
        os.makedirs(self.extracted_dir)
        return self

    def _check_safe_to_wipe(self):
        # The output dir is deleted wholesale on every run; refuse locations
        # where that would destroy data the script does not own.
        unsafe = {
            os.path.abspath(os.sep): "filesystem root",
            os.path.abspath(os.path.expanduser("~")): "home directory",
            os.path.abspath(os.getcwd()): "current working directory",
        }
        if self.root in unsafe:
            raise ValueError(
                f"Refusing to use {unsafe[self.root]} as output directory: "
                f"'{self.root}' would be deleted"
            )
        if os.path.exists(self.root) and not os.path.isdir(self.root):
            raise ValueError(f"Output path exists and is not a directory: '{self.root}'")

    @property
    def extracted_dir(self):
        return os.path.join(self.root, EXTRACTED_DIR_NAME)

    def path(self, name):
        return os.path.join(self.root, name)

    def extracted_event_path(self, index, kind):
        return os.path.join(self.extracted_dir, f"event_{index:02d}_{kind}.log")


# ========================================================================
# [S3] Log Extractor (Stage 1)
# ========================================================================

class _EventDraft:
    """An event being collected while scanning the raw log."""

    def __init__(self, index, kind, command, start_line):
        self.index = index
        self.kind = kind
        self.command = command
        self.start_line = start_line
        self.end_line = start_line
        self.time_start = None
        self.time_end = None
        self.saw_kernel_banner = False
        self.lines = []

    def add(self, line_no, text, timestamp):
        self.lines.append(text)
        self.end_line = line_no
        if timestamp:
            if self.time_start is None:
                self.time_start = timestamp
            self.time_end = timestamp


def _write_event(draft, output, end_reason):
    while draft.lines and not draft.lines[-1].strip():
        draft.lines.pop()
    path = output.extracted_event_path(draft.index, draft.kind)
    with open(path, "w", encoding="utf-8") as cleaned:
        cleaned.write("\n".join(draft.lines) + "\n")
    return ExtractedEvent(
        index=draft.index,
        kind=draft.kind,
        command=draft.command,
        time_start=draft.time_start,
        time_end=draft.time_end,
        source_line_start=draft.start_line,
        source_line_end=draft.end_line,
        extracted_path=path,
        end_reason=end_reason,
    )


def _clean_line(raw_line):
    """Decode a raw log line, split off the timestamp prefix."""
    text = raw_line.decode("utf-8", errors="replace").rstrip("\r\n")
    match = TIMESTAMP_RE.match(text)
    if match:
        return text[match.end():], match.group(1)
    return text, None


def extract_events(logfile, output):
    """Split the raw log into cleaned per-event files under output/extracted/.

    Detects alloc-fail and heapinfo events, strips timestamp prefixes, and
    writes each event to output/extracted/event_NN_<kind>.log. Returns the
    list of ExtractedEvent in order of appearance.
    """
    events = []
    current = None
    next_index = 1
    # Recent lines outside any event: pulled in as context when an
    # alloc-fail marker appears right after its mm_malloc error lines.
    lookback = deque(maxlen=3)

    def close(reason):
        nonlocal current
        if current is not None:
            events.append(_write_event(current, output, reason))
            current = None

    def open_event(kind, command, start_line):
        nonlocal current, next_index
        current = _EventDraft(next_index, kind, command, start_line)
        next_index += 1
        return current

    with open(logfile, "rb") as raw:
        for line_no, raw_line in enumerate(raw, 1):
            text, timestamp = _clean_line(raw_line)

            fail_matches = list(FAIL_START_RE.finditer(text))
            if fail_matches:
                # The marker may sit mid-line, glued to unrelated output
                # (and, pathologically, more than once per line).
                head = text[:fail_matches[0].start()]
                had_current = current is not None
                if had_current:
                    if head.strip():
                        current.add(line_no, head, timestamp)
                    close("next_event")
                for index, match in enumerate(fail_matches):
                    if index > 0:
                        close("next_event")
                    open_event(EVENT_ALLOC_FAIL, None, line_no)
                    if index == 0 and not had_current:
                        # Pull in the mm_malloc error context: either the
                        # glued fragment on this very line, or the lines
                        # immediately preceding the marker.
                        if FAIL_CONTEXT_RE.search(head):
                            current.add(line_no, head, timestamp)
                        else:
                            context = []
                            expected = line_no
                            for back_no, back_text, back_ts in reversed(lookback):
                                if (back_no == expected - 1
                                        and FAIL_CONTEXT_RE.search(back_text)):
                                    context.append((back_no, back_text, back_ts))
                                    expected = back_no
                                else:
                                    break
                            for back_no, back_text, back_ts in reversed(context):
                                current.add(back_no, back_text, back_ts)
                                current.start_line = min(current.start_line, back_no)
                    end = (fail_matches[index + 1].start()
                           if index + 1 < len(fail_matches) else len(text))
                    current.add(line_no, text[match.start():end], timestamp)
                continue

            tash_heapinfo = TASH_HEAPINFO_RE.search(text)
            if tash_heapinfo:
                close("next_event")
                open_event(EVENT_HEAPINFO, tash_heapinfo.group(1), line_no)
                current.add(line_no, text, timestamp)
                continue

            if KERNEL_BANNER_RE.search(text):
                # Each heapinfo run opens with exactly one of these banners.
                # A second banner inside a heapinfo event means a new run
                # whose TASH echo was lost. Inside an alloc-fail event the
                # kernel heap section is part of the dump: do not split.
                if current is None:
                    open_event(EVENT_HEAPINFO, None, line_no)
                elif (current.kind == EVENT_HEAPINFO
                      and current.saw_kernel_banner):
                    close("next_event")
                    open_event(EVENT_HEAPINFO, None, line_no)
                current.saw_kernel_banner = True
                current.add(line_no, text, timestamp)
                continue

            if REBOOT_RE.search(text):
                close("reboot")
                lookback.append((line_no, text, timestamp))
                continue

            if current is not None and TASH_PROMPT_RE.search(text):
                close("tash_prompt")
                lookback.append((line_no, text, timestamp))
                continue

            if current is not None:
                current.add(line_no, text, timestamp)
            else:
                lookback.append((line_no, text, timestamp))

    close("eof")
    return events


# ========================================================================
# [S4] Section Parsers (Stage 2)
# ========================================================================

# --- section header / field regexes -------------------------------------
APP_SUMMARY_BANNER_RE = re.compile(
    r"Summary of current app \(([^)]+)\) heap memory usage"
)
KERNEL_SUMMARY_BANNER_RE = re.compile(r"Summary of Kernel heap memory usage")
HEAPINFO_BANNER_RE = re.compile(r"^\s*(\S+) HEAP INFORMATION\s*$")
CORRUPTION_PREFIX_RE = re.compile(r"mm_check_heap_corruption:")
CORRUPTION_RANGE_RE = re.compile(
    r"Heap start = (0x[0-9a-fA-F]+) end = (0x[0-9a-fA-F]+)"
)
REGION_RE = re.compile(
    r"REGION\s+#(\d+)\s+Start=(0x(?:0x)?[0-9a-fA-F]+),\s+"
    r"End=(0x(?:0x)?[0-9a-fA-F]+),\s+Size=(\d+)"
)
TABLE_HEADER_RE = re.compile(r"MemAddr\s*\|\s*Size\s*\|\s*Status\s*\|")
TABLE_TERMINATOR_RE = re.compile(r"\*\* PID\(S\) in Pid column")
SUMMARY_USAGE_RE = re.compile(r"Summary of Heap Usages")

REQUESTED_RE = re.compile(r"requested size\s+(\d+)")
CALLER_RE = re.compile(r"caller address\s*=\s*(0x[0-9a-fA-F ]+)")
LARGEST_FREE_RE = re.compile(r"largest free size\s*:\s*(\d+)")
TOTAL_FREE_RE = re.compile(r"total free size\s*:\s*(\d+)")

BLOCK_ROW_RE = re.compile(
    r"^\s*(0x[0-9a-fA-F]+)\s*\|\s*(\d+)\s*\|\s*([AF])\s*\|"
    r"([^|]*)\|\s*([^|]*)\|\s*$"
)
ROW_CANDIDATE_RE = re.compile(r"^\s*0x[0-9a-fA-F]+\s*\|")
ADDRESS_RE = re.compile(r"0x[0-9a-fA-F]{8}")
TABLE_RULE_RE = re.compile(r"^[-\s|]+$")

SUM_TOTAL_RE = re.compile(r"^Total\s*:\s*(\d+)")
SUM_ALLOC_RE = re.compile(
    r"Allocated \(Current / Peak\)\s*:\s*(\d+)\s*\(\d+%\)\s*/\s*(\d+)"
)
SUM_FREE_RE = re.compile(r"Free \(Current\)\s*:\s*(\d+)")
SUM_RESERVED_RE = re.compile(r"Reserved\s*:\s*(\d+)")

DETAILS_RE = re.compile(r"Details of Heap Usages")
DET_FREE_NODES_RE = re.compile(r"Number of Free Node\s*:\s*(\d+)")
DET_LARGEST_RE = re.compile(r"Largest Free Node Size\s*:\s*(\d+)")
DET_DEAD_RE = re.compile(r"by Dead Threads \(\*\) \(1\)\s*:\s*(\d+)")
DET_STACK_RE = re.compile(r'Sum of "STACK"\(\*\*\) \(2\)\s*:\s*(\d+)')
DET_CURR_RE = re.compile(r'Sum of "CURR_HEAP" \(3\)\s*:\s*(\d+)')

NODELIST_HEADER_RE = re.compile(r"Available fragmented memory segments")
NODELIST_ROW_RE = re.compile(
    r"Nodelist\[(\d+)\] ranging \[(\d+), (\d+)\] : num (\d+), size (\d+)"
)

DEAD_THREADS_RE = re.compile(r"<\s*by Dead Threads\s*>")
DEAD_ROW_RE = re.compile(r"^\s*(\d+)\s*\|\s*(\d+)\s*$")
ALIVE_THREADS_RE = re.compile(r"<\s*by Alive Threads\s*>")
ALIVE_ROW_RE = re.compile(
    r"^\s*(\d+)\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*(-?\d+)\s*\|"
    r"\s*(-?\d+)\s*\|\s*(\S+)\s*\|\s*(.*?)\s*$"
)

ASSERTION_HEADER_RE = re.compile(r"Assertion details")
ASSERT_FAILED_RE = re.compile(
    r"Assertion failed .*?at file:\s*(\S+)\s+line\s+(\d+)"
    r"(?:\s+task:\s*(.+?))?(?:\s+pid:\s*(\d+))?\s*$"
)
ASSERT_PC_RE = re.compile(r"Assert location \(PC\)\s*:\s*(0x[0-9a-fA-F]+)")

STACK_HEADER_RE = re.compile(r"Asserted task's stack details")
STACK_SP_RE = re.compile(r"Current SP is .*SP:\s*([0-9a-fA-F]+)")
STACK_BASE_RE = re.compile(r"print_stack_dump:\s*base:\s*([0-9a-fA-F]+)")
STACK_SIZE_RE = re.compile(r"print_stack_dump:\s*size:\s*([0-9a-fA-F]+)")
STACK_USED_RE = re.compile(r"print_stack_dump:\s*used:\s*([0-9a-fA-F]+)")
STACKDUMP_PREFIX_RE = re.compile(r"^up_stackdump:")
STACKDUMP_ROW_RE = re.compile(
    r"^up_stackdump:\s*[0-9a-fA-F]+:(?:\s+[0-9a-fA-Fx]{8}){1,8}\s*$"
)

TCB_HEADER_RE = re.compile(r"Asserted task's TCB info")
TCB_FIELD_RE = re.compile(r"task_show_tcbinfo:\s*(\S[^:]*?)\s*:\s*(.*)$")

TASK_LIST_HEADER_RE = re.compile(r"NAME\s*\|\s*PID\s*\|\s*PRI\s*\|")
TASK_ROW_RE = re.compile(
    r"^\s*(.+?)\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*(\d+)\s*/\s*(\d+)\s*\|"
    r"\s*(0x[0-9a-fA-F]+)\s*\|\s*(0x[0-9a-fA-F]+)\s*\|\s*(\d+)\s*$"
)

LOADING_INFO_RE = re.compile(
    r"elf_show_all_bin_section_addr:\s*\[(\w+)\] Text Addr\s*:\s*"
    r"(0x[0-9a-fA-F]+), Text Size\s*:\s*(\d+)"
)

SEPARATOR_RE = re.compile(r"^\s*(?:mm_manage_alloc_fail\w*:\s*)?[*=#|-]{4,}\s*$")


# --- small parsing helpers (ported from heapfailanalyzer.py) ------------

def normalize_address(value):
    value = value.replace(" ", "").lower()
    while value.startswith("0x0x"):
        value = "0x" + value[4:]
    return value


def parse_owner(value):
    value = value.strip()
    if not re.fullmatch(r"0x\s*[0-9a-fA-F]+", value):
        return None
    return normalize_address(value)


def parse_pid(value):
    match = re.fullmatch(r"(\d+)\s*(?:\(\s*S\s*\))?", value.strip())
    if not match:
        return None, False
    return match.group(1), "(S)" in value.replace(" ", "")


class _LineStream:
    """Line iterator with one-line lookahead for section-boundary handling."""

    def __init__(self, lines):
        self._lines = lines
        self._pos = 0

    def peek(self):
        return self._lines[self._pos] if self._pos < len(self._lines) else None

    def next(self):
        line = self.peek()
        if line is not None:
            self._pos += 1
        return line


class _ParseContext:
    """Tracks which heap snapshot subsequent sections belong to."""

    def __init__(self, event):
        self.event = event
        self.heap = None
        self.region = None

    def switch_heap(self, name):
        self.region = None
        for heap in self.event.heaps:
            if heap.name == name:
                self.heap = heap
                return heap
        self.heap = HeapSnapshot(name=name)
        self.event.heaps.append(self.heap)
        return self.heap

    def ensure_heap(self):
        if self.heap is None:
            self.switch_heap("unknown")
        return self.heap


def _section_match(line):
    for regex, parser in SECTION_REGISTRY:
        match = regex.search(line)
        if match:
            return match, parser
    return None


# --- section parsers -----------------------------------------------------
# Each parser gets (stream, header line, header match, context). It consumes
# its own lines and stops before the next section header, which stays in the
# stream for the main loop.

def _parse_fail_info(stream, line, match, ctx):
    del line
    if ctx.event.fail_info is not None:
        return
    info = FailInfo(heap_kind=match.group(1))
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        field = REQUESTED_RE.search(ahead)
        if field:
            info.requested_size = int(field.group(1))
        field = CALLER_RE.search(ahead)
        if field:
            info.caller_address = normalize_address(field.group(1))
        field = LARGEST_FREE_RE.search(ahead)
        if field:
            info.largest_free_size = int(field.group(1))
        field = TOTAL_FREE_RE.search(ahead)
        if field:
            info.total_free_size = int(field.group(1))
        stream.next()
    ctx.event.fail_info = info


def _parse_app_banner(stream, line, match, ctx):
    del stream, line
    ctx.switch_heap(match.group(1))


def _parse_kernel_banner(stream, line, match, ctx):
    del stream, line, match
    ctx.switch_heap("kernel")


def _parse_heapinfo_banner(stream, line, match, ctx):
    del stream, line
    name = match.group(1)
    ctx.switch_heap("kernel" if name.upper() == "KERNEL" else name)


def _parse_corruption(stream, line, match, ctx):
    del stream, match
    heap = ctx.heap
    if heap is None:
        # Corruption re-checks in the assert tail run outside any heap
        # banner; attributing them to the last-seen heap would clobber it.
        return
    range_match = CORRUPTION_RANGE_RE.search(line)
    if range_match and heap.heap_range is None:
        heap.heap_range = (range_match.group(1), range_match.group(2))
    if heap.corruption is None:
        if "No heap corruption detected" in line:
            heap.corruption = False
        elif "corruption detected" in line:
            heap.corruption = True


def _parse_region(stream, line, match, ctx):
    del stream, line
    heap = ctx.ensure_heap()
    region = Region(
        number=int(match.group(1)),
        start=normalize_address(match.group(2)),
        end=normalize_address(match.group(3)),
        size=int(match.group(4)),
    )
    heap.regions.append(region)
    ctx.region = region


def _make_block(row, region_no, region=None):
    address = normalize_address(row.group(1))
    if not ADDRESS_RE.fullmatch(address):
        return None
    size = int(row.group(2))
    if region is not None:
        region_start = int(region.start, 16)
        region_end = int(region.end, 16)
        block_start = int(address, 16)
        if block_start < region_start or block_start + size > region_end:
            return None
    status = row.group(3)
    owner = parse_owner(row.group(4))
    pid, is_stack = parse_pid(row.group(5))
    if status == "A" and (owner is None or pid is None):
        return None
    if status == "F" and (row.group(4).strip() or row.group(5).strip()):
        return None
    return Block(
        region=region_no,
        address=address,
        size=size,
        status=status,
        owner=owner,
        pid=pid,
        is_stack=is_stack if status == "A" else False,
        category="stack" if status == "A" and is_stack else (
            "heap" if status == "A" else "free"
        ),
    )


def _parse_dump_table(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    region_no = ctx.region.number if ctx.region else 0
    diagnostics = ctx.event.diagnostics
    clean_end = False
    saw_partial_row = False

    while True:
        ahead = stream.peek()
        if ahead is None:
            break
        if TABLE_TERMINATOR_RE.search(ahead):
            stream.next()
            clean_end = True
            break
        if _section_match(ahead):
            break
        if (not ahead.strip() or TABLE_RULE_RE.match(ahead)
                or SEPARATOR_RE.match(ahead)):
            stream.next()
            continue
        row = BLOCK_ROW_RE.match(ahead)
        stream.next()
        if row:
            block = _make_block(row, region_no, ctx.region)
            if block:
                heap.blocks.append(block)
            else:
                diagnostics.malformed_rows += 1
        else:
            # Serial logs corrupt table rows in several observed ways: a
            # row cut short ("0x605c6810 |     "), a row with its middle
            # eaten ("0  A    | ... |"), or another log stream interleaved
            # character-by-character. The table only ends at a terminator,
            # a section header, or EOF; everything else is a damaged row.
            diagnostics.malformed_rows += 1
            saw_partial_row |= bool(ROW_CANDIDATE_RE.match(ahead))

    ahead = stream.peek()
    if saw_partial_row or (ahead is None and not clean_end):
        heap.blocks_truncated = True
        diagnostics.mark_truncated(f"dump_table:{heap.name}")


def _parse_summary(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    summary = HeapSummary()
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        matched = False
        field = SUM_TOTAL_RE.search(ahead)
        if field:
            summary.total = int(field.group(1))
            matched = True
        field = SUM_ALLOC_RE.search(ahead)
        if field:
            summary.alloc_current = int(field.group(1))
            summary.alloc_peak = int(field.group(2))
            matched = True
        field = SUM_FREE_RE.search(ahead)
        if field:
            summary.free_current = int(field.group(1))
            matched = True
        field = SUM_RESERVED_RE.search(ahead)
        if field:
            summary.reserved = int(field.group(1))
            matched = True
        if not matched and ahead.strip() and not SEPARATOR_RE.match(ahead):
            break
        stream.next()
    heap.summary = summary


def _parse_details(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    details = HeapDetails()
    fields = (
        (DET_FREE_NODES_RE, "free_node_count"),
        (DET_LARGEST_RE, "largest_free_node"),
        (DET_DEAD_RE, "dead_thread_alloc"),
        (DET_STACK_RE, "stack_sum"),
        (DET_CURR_RE, "curr_heap_sum"),
    )
    # The details body interleaves labels and notes with the value lines,
    # so consume everything up to the next section header.
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        for regex, attribute in fields:
            found = regex.search(ahead)
            if found:
                setattr(details, attribute, int(found.group(1)))
        stream.next()
    heap.details = details


def _parse_nodelist(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        row = NODELIST_ROW_RE.search(ahead)
        if row:
            heap.nodelist.append(NodeBucket(
                range_min=int(row.group(2)),
                range_max=int(row.group(3)),
                num=int(row.group(4)),
                size=int(row.group(5)),
            ))
        elif ahead.strip():
            break
        stream.next()


def _parse_dead_threads(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        row = DEAD_ROW_RE.match(ahead)
        if row:
            heap.dead_threads.append((int(row.group(1)), int(row.group(2))))
        elif ahead.strip() and not TABLE_RULE_RE.match(ahead) \
                and not SEPARATOR_RE.match(ahead) \
                and not re.search(r"Pid\s*\|\s*Size", ahead):
            break
        stream.next()


def _parse_alive_threads(stream, line, match, ctx):
    del line, match
    heap = ctx.ensure_heap()
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        row = ALIVE_ROW_RE.match(ahead)
        if row:
            name = row.group(7)
            heap.alive_threads.append(ThreadEntry(
                pid=int(row.group(1)),
                ppid=int(row.group(2)),
                stack=int(row.group(3)),
                curr_heap=int(row.group(4)),
                peak_heap=int(row.group(5)),
                bin=row.group(6),
                name=name[:-2] if name.endswith("()") else name,
            ))
            stream.next()
        elif (not ahead.strip() or TABLE_RULE_RE.match(ahead)
                or SEPARATOR_RE.match(ahead)
                or re.search(r"PID\s*\|\s*PPID", ahead)):
            stream.next()
        else:
            break


def _ensure_assertion(ctx):
    if ctx.event.assertion is None:
        ctx.event.assertion = AssertionInfo()
    return ctx.event.assertion


def _parse_assertion(stream, line, match, ctx):
    del line, match
    # Everything after the assert banner is system-wide (stack dump, task
    # list, corruption re-checks): stop attributing sections to a heap.
    ctx.heap = None
    info = _ensure_assertion(ctx)
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        failed = ASSERT_FAILED_RE.search(ahead)
        if failed:
            info.file = failed.group(1)
            info.line = int(failed.group(2))
            info.task = failed.group(3)
            info.pid = int(failed.group(4)) if failed.group(4) else None
        pc = ASSERT_PC_RE.search(ahead)
        if pc:
            info.pc = normalize_address(pc.group(1))
        stream.next()


def _parse_stack_details(stream, line, match, ctx):
    del line, match
    ctx.heap = None
    stack = StackInfo()
    fields = (
        (STACK_SP_RE, "sp", str),
        (STACK_BASE_RE, "base", str),
        (STACK_SIZE_RE, "size", lambda v: int(v, 16)),
        (STACK_USED_RE, "used", lambda v: int(v, 16)),
    )
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        for regex, attribute, convert in fields:
            found = regex.search(ahead)
            if found:
                setattr(stack, attribute, convert(found.group(1)))
        if (STACKDUMP_PREFIX_RE.match(ahead)
                and not STACKDUMP_ROW_RE.match(ahead)):
            # Pattern (3): the dump was cut mid-line and glued to whatever
            # came next (e.g. "... 604===========").
            stack.truncated = True
            ctx.event.diagnostics.mark_truncated("stack_dump")
        stream.next()
    _ensure_assertion(ctx).stack = stack


def _parse_tcb(stream, line, match, ctx):
    del line, match
    ctx.heap = None
    info = _ensure_assertion(ctx)
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        field = TCB_FIELD_RE.search(ahead)
        if field:
            info.tcb[field.group(1)] = field.group(2).strip()
        stream.next()


def _parse_task_list(stream, line, match, ctx):
    del line, match
    ctx.heap = None
    task_list = ctx.event.task_list or TaskList()
    while True:
        ahead = stream.peek()
        if ahead is None or _section_match(ahead):
            break
        row = TASK_ROW_RE.match(ahead)
        if row:
            task_list.entries.append(TaskEntry(
                name=row.group(1),
                pid=int(row.group(2)),
                priority=int(row.group(3)),
                stack_used=int(row.group(4)),
                stack_total=int(row.group(5)),
                stack_addr=row.group(6),
                tcb_addr=row.group(7),
                state=int(row.group(8)),
            ))
            stream.next()
        elif (not ahead.strip() or TABLE_RULE_RE.match(ahead)
                or SEPARATOR_RE.match(ahead)):
            stream.next()
        elif "|" in ahead:
            # Pattern (4): a task row cut short by the reboot.
            task_list.truncated = True
            ctx.event.diagnostics.mark_truncated("task_list")
            ctx.event.diagnostics.malformed_rows += 1
            stream.next()
        else:
            break
    ctx.event.task_list = task_list


def _parse_loading_info(stream, line, match, ctx):
    del stream, line
    ctx.event.loading_info.append(BinTextAddr(
        name=match.group(1),
        text_addr=normalize_address(match.group(2)),
        text_size=int(match.group(3)),
    ))


# Ordered: first matching regex wins. Narrow/specific patterns come before
# broader ones.
SECTION_REGISTRY = [
    (FAIL_START_RE, _parse_fail_info),
    (APP_SUMMARY_BANNER_RE, _parse_app_banner),
    (KERNEL_SUMMARY_BANNER_RE, _parse_kernel_banner),
    (HEAPINFO_BANNER_RE, _parse_heapinfo_banner),
    (CORRUPTION_PREFIX_RE, _parse_corruption),
    (REGION_RE, _parse_region),
    (TABLE_HEADER_RE, _parse_dump_table),
    (SUMMARY_USAGE_RE, _parse_summary),
    (DETAILS_RE, _parse_details),
    (NODELIST_HEADER_RE, _parse_nodelist),
    (DEAD_THREADS_RE, _parse_dead_threads),
    (ALIVE_THREADS_RE, _parse_alive_threads),
    (ASSERTION_HEADER_RE, _parse_assertion),
    (STACK_HEADER_RE, _parse_stack_details),
    (TCB_HEADER_RE, _parse_tcb),
    (TASK_LIST_HEADER_RE, _parse_task_list),
    (LOADING_INFO_RE, _parse_loading_info),
]


def parse_event(extracted):
    """Build a structured Event from one extracted event file."""
    event = Event(
        index=extracted.index,
        kind=extracted.kind,
        command=extracted.command,
        time_start=extracted.time_start,
        time_end=extracted.time_end,
        source_line_start=extracted.source_line_start,
        source_line_end=extracted.source_line_end,
        extracted_path=extracted.extracted_path,
        end_reason=extracted.end_reason,
    )
    with open(extracted.extracted_path, encoding="utf-8") as cleaned:
        stream = _LineStream(cleaned.read().splitlines())

    ctx = _ParseContext(event)
    while True:
        line = stream.next()
        if line is None:
            break
        matched = _section_match(line)
        if matched:
            match, parser = matched
            parser(stream, line, match, ctx)
        elif line.strip() and not SEPARATOR_RE.match(line):
            event.diagnostics.unknown_lines += 1
    return event


# ========================================================================
# [S5] Analyzer (Stage 3)
# ========================================================================

def _pid_directory(event):
    """pid(str) -> (task name, bin) from alive-thread tables and task list."""
    directory = {}
    for heap in event.heaps:
        for thread in heap.alive_threads:
            directory[str(thread.pid)] = (thread.name, thread.bin)
    if event.task_list:
        for task in event.task_list.entries:
            directory.setdefault(str(task.pid), (task.name, None))
    return directory


def make_histogram(blocks):
    """Bucket allocated blocks by size: 1-16, 17-32, 33-64, ... doubling."""
    allocated = [block for block in blocks if block.status == "A"]
    if not allocated:
        return []

    maximum = max(block.size for block in allocated)
    histogram = []
    lower, upper = 1, 16
    while lower <= maximum:
        upper = max(upper, lower)
        selected = [b for b in allocated if lower <= b.size <= upper]
        histogram.append({
            "label": f"{lower}-{upper}",
            "min": lower,
            "max": upper,
            "count": len(selected),
            "total_size": sum(b.size for b in selected),
            "heap_count": sum(not b.is_stack for b in selected),
            "stack_count": sum(b.is_stack for b in selected),
        })
        lower = upper + 1
        upper = max(upper * 2, 32)
    return histogram


def _find_continuity_gaps(heap):
    """Adjacent blocks in a region must tile the address space exactly."""
    gaps = []
    by_region = {}
    for block in heap.blocks:
        by_region.setdefault(block.region, []).append(block)
    for region_no, blocks in by_region.items():
        blocks = sorted(blocks, key=lambda b: int(b.address, 16))
        for previous, current in zip(blocks, blocks[1:]):
            expected = int(previous.address, 16) + previous.size
            actual = int(current.address, 16)
            if actual != expected:
                gaps.append(ContinuityGap(
                    region=region_no,
                    start=hex(expected),
                    end=hex(actual),
                    kind="gap" if actual > expected else "overlap",
                ))
    return gaps


def _fragmentation(heap, event):
    """(1 - largest free / total free) * 100; 0% = one contiguous free run.

    largest and total must come from the same snapshot source: either the
    details/summary pair, or the fail header of the failing heap.
    """
    largest = heap.details.largest_free_node if heap.details else None
    total = heap.summary.free_current if heap.summary else None
    if (largest is None or total is None) and event.fail_info \
            and event.heaps and heap is event.heaps[0]:
        largest = event.fail_info.largest_free_size
        total = event.fail_info.total_free_size
    if largest is None or not total:
        return None
    return (1.0 - largest / total) * 100.0


def _analyze_heap(heap, event, directory):
    derived = DerivedMetrics()
    derived.partial = heap.blocks_truncated
    derived.fragmentation_index = _fragmentation(heap, event)

    allocated = [b for b in heap.blocks if b.status == "A"]
    if heap.blocks:
        derived.total_allocated = sum(b.size for b in allocated)
        derived.observed_free = sum(
            b.size for b in heap.blocks if b.status == "F"
        )
        derived.histogram = make_histogram(heap.blocks)
        derived.continuity_gaps = _find_continuity_gaps(heap)

    largest = None
    for block in allocated:
        pid_name, pid_bin = directory.get(block.pid, (None, None))
        pid_data = derived.by_pid.setdefault(block.pid, {
            "count": 0, "total_allocated": 0, "heap_size": 0,
            "stack_size": 0, "name": pid_name, "bin": pid_bin,
        })
        pid_data["count"] += 1
        pid_data["total_allocated"] += block.size
        pid_data["stack_size" if block.is_stack else "heap_size"] += block.size

        owner_data = derived.by_owner.setdefault(block.owner, {
            "count": 0, "total_allocated": 0, "pids": [],
        })
        owner_data["count"] += 1
        owner_data["total_allocated"] += block.size
        if block.pid not in owner_data["pids"]:
            owner_data["pids"].append(block.pid)

        if largest is None or block.size > largest["size"]:
            largest = {
                "size": block.size, "pid": block.pid,
                "owner": block.owner, "address": block.address,
                "task": pid_name,
            }
    derived.largest_allocation = largest

    by_total = lambda item: item[1]["total_allocated"]
    derived.top_pids = [
        {"pid": pid, **data}
        for pid, data in sorted(derived.by_pid.items(), key=by_total,
                                reverse=True)[:5]
    ]
    derived.top_owners = [
        {"owner": owner, **data}
        for owner, data in sorted(derived.by_owner.items(), key=by_total,
                                  reverse=True)[:5]
    ]
    return derived


def analyze_event(event):
    """Compute derived metrics for every heap snapshot of the event."""
    directory = _pid_directory(event)
    for heap in event.heaps:
        heap.derived = _analyze_heap(heap, event, directory)


# ========================================================================
# [S6] Symbol Resolver (Stage 4)
# ========================================================================

ELF_NAMES = ("tinyara.axf", "common_dbg", "app1_dbg")


def _addr2line_result(function, location):
    resolved = function not in ("??", "") and location not in ("??:0", "??:?")
    return {
        "function": function or "??",
        "location": location or "??:0",
        "resolved": resolved,
    }


def _find_elf_files(bin_dir):
    script_bin = os.path.abspath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "..", "build", "output", "bin",
    ))
    directories = [bin_dir, script_bin,
                   os.path.abspath(os.path.join("build", "output", "bin"))]
    elf_paths = []
    for directory in directories:
        for name in ELF_NAMES:
            path = os.path.abspath(os.path.join(directory, name))
            if path not in elf_paths and os.path.isfile(path):
                elf_paths.append(path)
    return elf_paths


def _collect_addresses(result):
    addresses = set()
    for event in result.events:
        if event.fail_info and event.fail_info.caller_address:
            addresses.add(event.fail_info.caller_address)
        if event.assertion and event.assertion.pc:
            addresses.add(event.assertion.pc)
        for heap in event.heaps:
            for block in heap.blocks:
                if block.owner:
                    addresses.add(block.owner)
    return sorted(addresses)


def resolve_symbols(result, bin_dir, command, no_symbols=False):
    """Resolve owner/caller/PC addresses via addr2line, one batch per ELF.

    A resolved answer from a later ELF wins over "??" from an earlier one:
    kernel addresses live in tinyara.axf, app addresses in the *_dbg ELFs.
    """
    if no_symbols:
        result.symbols_skip_reason = "disabled by --no-symbols"
        return
    elf_paths = _find_elf_files(bin_dir)
    if not elf_paths:
        result.symbols_skip_reason = (
            f"no ELF binaries found ({'/'.join(ELF_NAMES)} under '{bin_dir}')"
        )
        return
    addresses = _collect_addresses(result)
    if not addresses:
        result.symbols_skip_reason = "no addresses to resolve"
        return

    symbols = {}
    for elf in elf_paths:
        try:
            completed = subprocess.run(
                [command, "-f", "-C", "-e", elf] + addresses,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
                check=False,
            )
        except OSError:
            result.symbols_skip_reason = f"'{command}' not available"
            return
        lines = completed.stdout.splitlines()
        for index, address in enumerate(addresses):
            function = lines[index * 2].strip() if index * 2 < len(lines) else "??"
            location = (lines[index * 2 + 1].strip()
                        if index * 2 + 1 < len(lines) else "??:0")
            entry = _addr2line_result(function, location)
            if address not in symbols or entry["resolved"]:
                entry["elf"] = os.path.basename(elf)
                symbols[address] = entry

    for event in result.events:
        if event.fail_info and event.fail_info.caller_address:
            event.fail_info.caller_symbol = symbols.get(
                event.fail_info.caller_address)
        if event.assertion and event.assertion.pc:
            event.assertion.pc_symbol = symbols.get(event.assertion.pc)
        for heap in event.heaps:
            for block in heap.blocks:
                if block.owner:
                    block.symbol = symbols.get(block.owner)
    result.symbols_resolved = any(
        entry["resolved"] for entry in symbols.values()
    )
    if not result.symbols_resolved:
        result.symbols_skip_reason = (
            "addr2line resolved no addresses (ELF/log mismatch?)"
        )


# ========================================================================
# [S7] Reporters (Stage 5)
# ========================================================================

def format_size(size_bytes):
    if size_bytes is None:
        return "N/A"
    size_bytes = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if size_bytes < 1024.0 or unit == "GB":
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0


def _size_pair(value):
    return f"{value} B ({format_size(value)})" if value is not None else "N/A"


def _pid_label(pid, name, bin_name):
    label = pid if pid is not None else "?"
    if name:
        label += f" {name}"
        if bin_name:
            label += f"[{bin_name}]"
    return label


def _render_heap(heap, out):
    heap_range = (f"{heap.heap_range[0]}..{heap.heap_range[1]}"
                  if heap.heap_range else "range unknown")
    corruption = {None: "not checked", True: "DETECTED!", False: "none"}
    out.append("")
    out.append(f"-- HEAP [{heap.name}] ({heap_range}, "
               f"corruption: {corruption[heap.corruption]}) --")

    if heap.summary:
        summary = heap.summary
        out.append(f"  Total    : {_size_pair(summary.total)}")
        out.append(f"  Allocated: {_size_pair(summary.alloc_current)}"
                   f"  (peak {_size_pair(summary.alloc_peak)})")
        out.append(f"  Free     : {_size_pair(summary.free_current)}"
                   f"   Reserved: {summary.reserved}")
    if heap.details:
        details = heap.details
        out.append(f"  Details  : free nodes {details.free_node_count}, "
                   f"largest free {_size_pair(details.largest_free_node)}")
        out.append(f"             dead-thread alloc {details.dead_thread_alloc}, "
                   f"alive stack {details.stack_sum}, "
                   f"alive heap {details.curr_heap_sum}")

    derived = heap.derived
    if derived is None:
        return
    if derived.fragmentation_index is not None:
        out.append(f"  Fragmentation index: {derived.fragmentation_index:.2f}%")
    if derived.total_allocated is not None:
        partial = "  (LOWER BOUND: dump truncated)" if derived.partial else ""
        out.append(f"  Dump observed: allocated {_size_pair(derived.total_allocated)}, "
                   f"free {_size_pair(derived.observed_free)}{partial}")
        if heap.summary and heap.summary.alloc_current is not None \
                and not derived.partial:
            diff = heap.summary.alloc_current - derived.total_allocated
            if diff:
                out.append(f"  Reported-vs-observed alloc difference: {diff} B "
                           "(heap metadata / lost rows)")
    if derived.continuity_gaps:
        kinds = [gap.kind for gap in derived.continuity_gaps]
        out.append(f"  Address continuity: {kinds.count('gap')} gap(s), "
                   f"{kinds.count('overlap')} overlap(s) in dump")

    if derived.by_pid:
        by_total = sorted(derived.by_pid.items(),
                          key=lambda item: item[1]["total_allocated"],
                          reverse=True)
        tasks = {}
        for pid, data in by_total:
            task = data["name"] or "?"
            if data["bin"]:
                task += f" [{data['bin']}]"
            tasks[pid] = task
        # Never truncate task names: size the column to the longest one.
        task_width = max([len(task) for task in tasks.values()] + [4])
        total = derived.total_allocated or 0
        out.append("")
        out.append(f"  {'PID':<6} {'Task':<{task_width}} {'Count':>6} "
                   f"{'Total':>10} {'Heap':>10} {'Stack':>10} {'%':>7}")
        for pid, data in by_total:
            percent = data["total_allocated"] * 100 / total if total else 0
            out.append(f"  {pid or '?':<6} {tasks[pid]:<{task_width}} "
                       f"{data['count']:>6} {data['total_allocated']:>10} "
                       f"{data['heap_size']:>10} {data['stack_size']:>10} "
                       f"{percent:>6.2f}%")

    if derived.by_owner:
        symbol_by_owner = {}
        for block in heap.blocks:
            if (block.owner and block.symbol
                    and block.owner not in symbol_by_owner):
                symbol_by_owner[block.owner] = block.symbol
        out.append("")
        out.append(f"  {'Owner':<16} {'Count':>6} {'Total':>10}  "
                   f"PIDs / Code")
        # Every owner, every PID -- the report must never hide data
        # behind an ellipsis. When addr2line ran, its result is shown
        # verbatim (including "?? ??:0"); without a binary, owner only.
        for owner, data in sorted(derived.by_owner.items(),
                                  key=lambda item: item[1]["total_allocated"],
                                  reverse=True):
            pids = ", ".join(sorted(data["pids"], key=int))
            symbol = symbol_by_owner.get(owner)
            code = ""
            if symbol:
                code = (f"  <{symbol.get('function', '??')} "
                        f"{symbol.get('location', '??:0')}>")
            out.append(f"  {owner or '?':<16} {data['count']:>6} "
                       f"{data['total_allocated']:>10}  {pids}{code}")

    if derived.histogram:
        out.append("")
        out.append(f"  {'Size range':<20} {'Count':>6} {'Heap':>6} "
                   f"{'Stack':>6} {'Total':>12}")
        for bucket in derived.histogram:
            out.append(f"  {bucket['label']:<20} {bucket['count']:>6} "
                       f"{bucket['heap_count']:>6} {bucket['stack_count']:>6} "
                       f"{bucket['total_size']:>12}")

    largest = derived.largest_allocation
    if largest:
        task = f" task {largest['task']}" if largest.get("task") else ""
        out.append("")
        out.append(f"  Largest allocation: {_size_pair(largest['size'])} at "
                   f"{largest['address']} (PID {largest['pid']}{task}, "
                   f"owner {largest['owner']})")

    if heap.dead_threads:
        dead = ", ".join(f"pid {pid}: {size} B"
                         for pid, size in heap.dead_threads)
        out.append(f"  Dead-thread allocations (leak candidates): {dead}")

    live_buckets = [b for b in heap.nodelist if b.num]
    if live_buckets:
        buckets = ", ".join(f"[{b.range_min}-{b.range_max}] x{b.num} ({b.size} B)"
                            for b in live_buckets)
        out.append(f"  Free-node distribution: {buckets}")


def _render_event(event, out):
    out.append("")
    out.append("=" * 72)
    time_range = event.time_start or "time unknown"
    if event.time_end and event.time_end != event.time_start:
        time_range += f" ~ {event.time_end}"
    title = f"EVENT #{event.index}: {event.kind}"
    if event.command:
        title += f"  (command: {event.command})"
    out.append(title)
    out.append(f"  {time_range} | source lines "
               f"{event.source_line_start}-{event.source_line_end} | "
               f"ended by {event.end_reason}")
    out.append("=" * 72)

    diagnostics = event.diagnostics
    if diagnostics.truncated_sections:
        out.append(f"  (!) TRUNCATED sections: "
                   f"{', '.join(diagnostics.truncated_sections)}"
                   " -- affected numbers are lower bounds")
    if diagnostics.malformed_rows:
        out.append(f"  (!) Malformed/corrupted rows skipped: "
                   f"{diagnostics.malformed_rows}")

    info = event.fail_info
    if info:
        out.append("")
        out.append(f"[FAIL] allocation from {info.heap_kind or '?'} heap")
        out.append(f"  requested   : {_size_pair(info.requested_size)}")
        caller = info.caller_address or "N/A"
        if info.caller_symbol and info.caller_symbol.get("resolved"):
            caller += (f"  {info.caller_symbol['function']} "
                       f"({info.caller_symbol['location']})")
        out.append(f"  caller      : {caller}")
        out.append(f"  largest free: {_size_pair(info.largest_free_size)}")
        out.append(f"  total free  : {_size_pair(info.total_free_size)}")
        if info.requested_size and info.largest_free_size is not None:
            if info.requested_size > info.largest_free_size:
                cause = ("no free node is large enough"
                         if info.total_free_size
                         and info.requested_size <= info.total_free_size
                         else "request exceeds total free memory")
                out.append(f"  => {cause}")

    for heap in event.heaps:
        _render_heap(heap, out)

    assertion = event.assertion
    if assertion:
        out.append("")
        out.append(f"[ASSERT] {assertion.file}:{assertion.line} "
                   f"task {assertion.task} (pid {assertion.pid})")
        pc = assertion.pc or "N/A"
        if assertion.pc_symbol and assertion.pc_symbol.get("resolved"):
            pc += (f"  {assertion.pc_symbol['function']} "
                   f"({assertion.pc_symbol['location']})")
        out.append(f"  PC: {pc}")
        stack = assertion.stack
        if stack:
            truncated = "  (dump truncated)" if stack.truncated else ""
            out.append(f"  Stack: base {stack.base}, size {stack.size}, "
                       f"used {stack.used}{truncated}")

    if event.task_list and event.task_list.entries:
        note = " (TRUNCATED)" if event.task_list.truncated else ""
        out.append(f"[TASKS] {len(event.task_list.entries)} tasks captured{note}")
        heavy = sorted(event.task_list.entries,
                       key=lambda task: task.stack_used, reverse=True)[:3]
        for task in heavy:
            out.append(f"  top stack use: {task.name} (pid {task.pid}) "
                       f"{task.stack_used}/{task.stack_total}")

    if event.loading_info:
        bins = ", ".join(f"{b.name}@{b.text_addr}(+{b.text_size})"
                         for b in event.loading_info)
        out.append(f"[BINARIES] {bins}")


def render_console_report(result):
    """Return the human-readable report as a string (stdout + report.txt)."""
    lines = []
    lines.append("===== TizenRT HEAP LOG ANALYSIS =====")
    lines.append(f"Source: {result.source.path} ({format_size(result.source.size)})")
    lines.append(f"Events found: {result.source.total_events}")
    if result.symbols_skip_reason:
        lines.append(f"Symbols: skipped ({result.symbols_skip_reason})")

    if not result.events:
        lines.append("")
        lines.append("No alloc-fail or heapinfo events detected in the log.")
        return "\n".join(lines) + "\n"

    lines.append("")
    lines.append("--- EVENT LIST ---")
    header = f"{'#':<4} {'Kind':<12} {'Time':<26} {'Lines':<15} {'Truncated'}"
    lines.append(header)
    for event in result.events:
        time_start = event.time_start or "N/A"
        line_range = f"{event.source_line_start}-{event.source_line_end}"
        truncated = ", ".join(event.diagnostics.truncated_sections) or "-"
        lines.append(
            f"{event.index:<4} {event.kind:<12} {time_start:<26} "
            f"{line_range:<15} {truncated}"
        )

    for event in result.events:
        _render_event(event, lines)
    return "\n".join(lines) + "\n"


def report_console(result, output):
    text = render_console_report(result)
    sys.stdout.write(text)
    with open(output.path(REPORT_TXT_NAME), "w", encoding="utf-8") as report:
        report.write(text)


def _json_default(value):
    if isinstance(value, Enum):
        return value.value
    raise TypeError(f"not JSON serializable: {type(value)!r}")


def report_json(result, output):
    payload = asdict(result)
    with open(output.path(ANALYSIS_JSON_NAME), "w", encoding="utf-8") as dump:
        json.dump(payload, dump, indent=2, default=_json_default)


def _excel_heap_sheets(workbook, event, heap, prefix):
    derived = heap.derived or DerivedMetrics()
    tag = f"{prefix}{heap.name[:10]}"

    if derived.by_pid:
        sheet = workbook.create_sheet(f"{tag}_PID"[:31])
        sheet.append(["PID", "Task", "Bin", "Count", "Total Allocated",
                      "Heap", "Stack", "% of Allocated"])
        total = derived.total_allocated or 0
        for pid, data in sorted(derived.by_pid.items(),
                                key=lambda item: item[1]["total_allocated"],
                                reverse=True):
            sheet.append([
                pid, data["name"] or "", data["bin"] or "", data["count"],
                data["total_allocated"], data["heap_size"], data["stack_size"],
                data["total_allocated"] * 100 / total if total else 0,
            ])

    if derived.by_owner:
        symbol_by_owner = {}
        for block in heap.blocks:
            if block.owner and block.symbol and block.owner not in symbol_by_owner:
                symbol_by_owner[block.owner] = block.symbol
        sheet = workbook.create_sheet(f"{tag}_Owner"[:31])
        sheet.append(["Owner", "Count", "Total Allocated", "% of Allocated",
                      "PIDs", "Function", "Location", "ELF"])
        total = derived.total_allocated or 0
        for owner, data in sorted(derived.by_owner.items(),
                                  key=lambda item: item[1]["total_allocated"],
                                  reverse=True):
            symbol = symbol_by_owner.get(owner) or {}
            sheet.append([
                owner, data["count"], data["total_allocated"],
                data["total_allocated"] * 100 / total if total else 0,
                ", ".join(sorted(data["pids"], key=int)),
                symbol.get("function", ""),
                symbol.get("location", ""), symbol.get("elf", ""),
            ])

    if derived.histogram:
        sheet = workbook.create_sheet(f"{tag}_Histogram"[:31])
        sheet.append(["Range", "Count", "Heap Count", "Stack Count",
                      "Total Size"])
        for bucket in derived.histogram:
            sheet.append([bucket["label"], bucket["count"],
                          bucket["heap_count"], bucket["stack_count"],
                          bucket["total_size"]])

    if heap.blocks:
        sheet = workbook.create_sheet(f"{tag}_Blocks"[:31])
        sheet.append(["Region", "Address", "Size", "Status", "Category",
                      "Owner", "PID", "Task", "Function", "Location", "ELF"])
        directory = _pid_directory(event)
        for block in heap.blocks:
            symbol = block.symbol or {}
            task = directory.get(block.pid, ("", None))[0] or ""
            sheet.append([
                block.region, block.address, block.size, block.status,
                block.category, block.owner or "", block.pid or "", task,
                symbol.get("function", ""), symbol.get("location", ""),
                symbol.get("elf", ""),
            ])


def _excel_event_sheets(workbook, event, prefix):
    summary_sheet = workbook.create_sheet(f"{prefix}Summary"[:31])
    summary_sheet.append(["Section", "Metric", "Value", "Human Readable"])
    if event.fail_info:
        info = event.fail_info
        for metric, value in (
            ("Failed heap", info.heap_kind),
            ("Requested size", info.requested_size),
            ("Caller address", info.caller_address),
            ("Caller function",
             (info.caller_symbol or {}).get("function", "")),
            ("Largest free size", info.largest_free_size),
            ("Total free size", info.total_free_size),
        ):
            readable = (format_size(value)
                        if isinstance(value, int) else "")
            summary_sheet.append(["FAIL", metric, value, readable])
    for heap in event.heaps:
        derived = heap.derived or DerivedMetrics()
        summary = heap.summary or HeapSummary()
        rows = [
            ("Heap range", f"{heap.heap_range[0]}..{heap.heap_range[1]}"
             if heap.heap_range else ""),
            ("Corruption", heap.corruption),
            ("Total", summary.total),
            ("Allocated (current)", summary.alloc_current),
            ("Allocated (peak)", summary.alloc_peak),
            ("Free (current)", summary.free_current),
            ("Reserved", summary.reserved),
            ("Dump observed allocated", derived.total_allocated),
            ("Dump observed free", derived.observed_free),
            ("Fragmentation index (%)", derived.fragmentation_index),
            ("Dump truncated", heap.blocks_truncated),
        ]
        for metric, value in rows:
            readable = format_size(value) if isinstance(value, int) else ""
            summary_sheet.append([f"HEAP {heap.name}", metric, value, readable])

    for heap in event.heaps:
        _excel_heap_sheets(workbook, event, heap, prefix)

    threads = [(heap.name, thread) for heap in event.heaps
               for thread in heap.alive_threads]
    dead = [(heap.name, pid, size) for heap in event.heaps
            for pid, size in heap.dead_threads]
    if threads or dead:
        sheet = workbook.create_sheet(f"{prefix}Threads"[:31])
        sheet.append(["Heap", "PID", "PPID", "Stack", "Curr Heap",
                      "Peak Heap", "Bin", "Name"])
        for heap_name, thread in threads:
            sheet.append([heap_name, thread.pid, thread.ppid, thread.stack,
                          thread.curr_heap, thread.peak_heap, thread.bin,
                          thread.name])
        if dead:
            sheet.append([])
            sheet.append(["Heap", "Dead PID", "Alive allocation"])
            for heap_name, pid, size in dead:
                sheet.append([heap_name, pid, size])

    if event.task_list and event.task_list.entries:
        sheet = workbook.create_sheet(f"{prefix}Tasks"[:31])
        sheet.append(["Name", "PID", "Priority", "Stack Used", "Stack Total",
                      "Stack Addr", "TCB Addr", "State"])
        for task in event.task_list.entries:
            sheet.append([task.name, task.pid, task.priority, task.stack_used,
                          task.stack_total, task.stack_addr, task.tcb_addr,
                          task.state])


def report_excel(result, output):
    """Write report.xlsx with one sheet group per event."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Warning: openpyxl not installed, skipping Excel export",
              file=sys.stderr)
        return

    workbook = Workbook()
    events_sheet = workbook.active
    events_sheet.title = "Events"
    events_sheet.append(["#", "Kind", "Command", "Time Start", "Time End",
                         "Source Lines", "End Reason", "Truncated Sections",
                         "Malformed Rows"])
    for event in result.events:
        events_sheet.append([
            event.index, event.kind, event.command or "",
            event.time_start or "", event.time_end or "",
            f"{event.source_line_start}-{event.source_line_end}",
            event.end_reason,
            ", ".join(event.diagnostics.truncated_sections),
            event.diagnostics.malformed_rows,
        ])

    for event in result.events:
        prefix = f"E{event.index}_" if len(result.events) > 1 else ""
        _excel_event_sheets(workbook, event, prefix)

    diagnostics_sheet = workbook.create_sheet("Diagnostics")
    diagnostics_sheet.append(["Event", "Metric", "Value"])
    diagnostics_sheet.append(["-", "Source file", result.source.path])
    diagnostics_sheet.append(["-", "File size", result.source.size])
    diagnostics_sheet.append(
        ["-", "Symbols", "resolved" if result.symbols_resolved
         else result.symbols_skip_reason or "n/a"])
    for event in result.events:
        diagnostics = event.diagnostics
        diagnostics_sheet.append(
            [event.index, "Truncated sections",
             ", ".join(diagnostics.truncated_sections) or "-"])
        diagnostics_sheet.append(
            [event.index, "Malformed rows", diagnostics.malformed_rows])
        diagnostics_sheet.append(
            [event.index, "Unknown lines", diagnostics.unknown_lines])
        for heap in event.heaps:
            if heap.derived and heap.derived.continuity_gaps:
                for gap in heap.derived.continuity_gaps:
                    diagnostics_sheet.append(
                        [event.index, f"Continuity {gap.kind} ({heap.name})",
                         f"region {gap.region}: {gap.start}..{gap.end}"])
    workbook.save(output.path(REPORT_XLSX_NAME))


def _display_blocks(heap, directory):
    """Blocks plus inferred gap fillers, as JSON-ready dicts for the map."""
    def as_dict(block, task):
        symbol = block.symbol or {}
        return {
            "region": block.region, "address": block.address,
            "size": block.size, "status": block.status,
            "category": block.category, "pid": block.pid,
            "task": task, "owner": block.owner,
            "function": symbol.get("function"),
            "location": symbol.get("location"),
            "inferred": block.inferred,
        }

    def gap(region_no, start, size):
        return {
            "region": region_no, "address": f"0x{start:08x}", "size": size,
            "status": "G", "category": "gap", "pid": None, "task": None,
            "owner": None, "function": None, "location": None,
            "inferred": True,
        }

    by_region = {}
    for block in heap.blocks:
        by_region.setdefault(block.region, []).append(block)

    display = []
    for region in heap.regions:
        cursor = int(region.start, 16)
        region_end = int(region.end, 16)
        for block in sorted(by_region.get(region.number, []),
                            key=lambda b: int(b.address, 16)):
            address = int(block.address, 16)
            if address > cursor:
                display.append(gap(region.number, cursor, address - cursor))
            task = directory.get(block.pid, (None, None))[0]
            display.append(as_dict(block, task))
            cursor = max(cursor, address + block.size)
        if cursor < region_end:
            display.append(gap(region.number, cursor, region_end - cursor))

    known = {region.number for region in heap.regions}
    for block in heap.blocks:
        if block.region not in known:
            display.append(as_dict(block,
                                   directory.get(block.pid, (None, None))[0]))
    return sorted(display,
                  key=lambda b: (b["region"], int(b["address"], 16)))


def _html_runs(heap, directory=None):
    """Merge address-contiguous blocks of the same category into runs.

    The map itself only shows heap/stack/free/unknown, but each run keeps
    its member allocations (address/size/pid/task/owner/symbol) so the
    detail panel can answer "who allocated how much here" on click.
    Gap inference happens first so the runs tile each region.
    """
    runs = []
    for block in _display_blocks(heap, directory or {}):
        category = ("unknown" if block["category"] == "gap"
                    else block["category"])
        member = None
        if block["status"] == "A":
            member = {
                "address": block["address"],
                "size": block["size"],
                "pid": block["pid"],
                "task": block["task"],
                "owner": block["owner"],
                "function": block["function"],
                "location": block["location"],
            }
        previous = runs[-1] if runs else None
        contiguous = (
            previous is not None
            and previous["region"] == block["region"]
            and previous["category"] == category
            and int(previous["address"], 16) + previous["size"]
                == int(block["address"], 16)
        )
        if contiguous:
            previous["size"] += block["size"]
            previous["count"] += 1
            if member:
                previous["blocks"].append(member)
        else:
            runs.append({
                "region": block["region"],
                "address": block["address"],
                "size": block["size"],
                "category": category,
                "count": 1,
                "blocks": [member] if member else [],
            })
    return runs


# Vertical strip layout: address-contiguous runs of the same category
# (heap / stack / free / unknown) are stacked top-to-bottom in address
# order. Row height steps by size bucket (<=16 B, <=32 B, <=64 B, ...,
# doubling like the histogram), NOT proportional to byte size -- so the
# rough size class is readable from the picture alone while a 32 B chunk
# stays visible next to a 5 MB one. Every row is tall enough for its
# one-line label.
HTML_TEMPLATE = """<!doctype html>
<meta charset="utf-8">
<title>TizenRT Heap Memory Map</title>
<style>
/* Role tokens. Series colors validated for CVD separation and contrast
   (light: #2a78d6/#eb6834/#199e70, dark steps likewise); "unknown" is a
   neutral non-data marker: gray hatch + always-on text label. */
:root {
  color-scheme: light;
  --surface: #fcfcfb;      --surface-alt: #f4f4f1;
  --border: #e3e2de;       --border-strong: #c9c8c2;
  --text-1: #0b0b0b;       --text-2: #52514e;       --text-3: #7c7b73;
  --heap: #2a78d6;         --stack: #eb6834;        --free: #199e70;
  --unknown-1: #d3d6da;    --unknown-2: #e7e9eb;    --unknown-text: #3f4246;
  --warn: #b3261e;         --on-series: #ffffff;
  --select-ring: #0b0b0b;
}
@media (prefers-color-scheme: dark) {
  :root {
    color-scheme: dark;
    --surface: #1a1a19;    --surface-alt: #232322;
    --border: #3a3a38;     --border-strong: #4c4c49;
    --text-1: #f4f4f0;     --text-2: #c3c2b7;       --text-3: #8f8e85;
    --heap: #3987e5;       --stack: #d95926;        --free: #199e70;
    --unknown-1: #3b3e41;  --unknown-2: #2e3032;    --unknown-text: #c3c2b7;
    --warn: #f28b82;       --on-series: #ffffff;
    --select-ring: #f4f4f0;
  }
}
* { box-sizing: border-box; }
html, body { height: 100%; }
body { margin: 0; display: flex; flex-direction: column;
       background: var(--surface); color: var(--text-1);
       font: 14px/1.45 system-ui, -apple-system, "Segoe UI", sans-serif; }
code, pre, .run, table { font-family: ui-monospace, SFMono-Regular, Menlo,
                         Consolas, monospace; }

/* ---- top bar ---- */
#top { flex: none; padding: 10px 16px 8px;
       border-bottom: 1px solid var(--border); }
#top .bar { display: flex; align-items: center; flex-wrap: wrap;
            gap: 8px 14px; }
#top h1 { font-size: 15px; font-weight: 700; margin: 0 8px 0 0; }
#top label { color: var(--text-2); font-size: 13px; }
#top select { font: 13px system-ui, sans-serif; color: var(--text-1);
              background: var(--surface); border: 1px solid var(--border-strong);
              border-radius: 6px; padding: 3px 6px; }
.legend { display: inline-flex; gap: 6px; align-items: center; }
.legend .key { display: inline-flex; align-items: center; gap: 5px;
               font-size: 12px; color: var(--text-2); }
.legend .swatch { width: 12px; height: 12px; border-radius: 3px; }
.swatch.heap { background: var(--heap); }
.swatch.stack { background: var(--stack); }
.swatch.free { background: var(--free); }
.swatch.unknown { background: repeating-linear-gradient(45deg,
  var(--unknown-1) 0 3px, var(--unknown-2) 3px 6px); }
#stats { margin-top: 7px; display: flex; flex-wrap: wrap; gap: 6px 8px;
         font-size: 12px; }
#stats .stat { padding: 2px 9px; border: 1px solid var(--border);
               border-radius: 999px; background: var(--surface-alt);
               color: var(--text-2); }
#stats .stat b { color: var(--text-1); }
.warn { color: var(--warn); font-weight: 600; }

/* ---- three panels ---- */
#main { flex: 1; display: flex; overflow: hidden; }
.panel { flex: 1 1 0; min-width: 0; display: flex; flex-direction: column; }
.panel-title { flex: none; padding: 7px 16px 5px; font-size: 11px;
               font-weight: 700; letter-spacing: .08em;
               text-transform: uppercase; color: var(--text-3);
               border-bottom: 1px solid var(--border);
               background: var(--surface-alt); }
.panel-body { flex: 1; overflow: auto; }
.divider { flex: none; width: 2px; background: var(--border-strong); }

/* ---- left: memory strip ---- */
#strip { padding: 6px 16px 60px; }
.region-header { position: sticky; top: 0; z-index: 1;
                 background: var(--surface); font-size: 12px;
                 font-weight: 600; color: var(--text-2);
                 padding: 8px 0 6px; }
.run { display: flex; align-items: center; padding: 0 8px;
       overflow: hidden; white-space: nowrap;
       font-size: 12px; line-height: 1; color: var(--on-series);
       border-radius: 4px; margin-bottom: 2px;   /* 2px surface gap */
       cursor: pointer; }
.run:hover { filter: brightness(1.12); }
.run:focus-visible { outline: 2px solid var(--select-ring);
                     outline-offset: 1px; }
.run.selected { outline: 2px solid var(--select-ring);
                outline-offset: -2px; box-shadow: inset 0 0 0 3px
                var(--surface); }
.run.heap    { background: var(--heap); }
.run.stack   { background: var(--stack); }
.run.free    { background: var(--free); font-weight: 600; }
.run.unknown { color: var(--unknown-text); font-weight: 600; background:
  repeating-linear-gradient(45deg, var(--unknown-1) 0 6px,
                            var(--unknown-2) 6px 12px); }

/* ---- middle: selection detail ---- */
#detail { padding: 12px 16px 60px; font-size: 13px; }
#detail .placeholder { color: var(--text-3); }
#detail h3 { margin: 0 0 4px; font-size: 14px; }
#detail .sub { color: var(--text-2); margin-bottom: 12px; }
#detail h4 { margin: 14px 0 6px; font-size: 11px; font-weight: 700;
             letter-spacing: .08em; text-transform: uppercase;
             color: var(--text-3); }
#detail table { border-collapse: collapse; width: 100%; font-size: 12px; }
#detail th, #detail td { padding: 4px 10px; text-align: right;
                         border-bottom: 1px solid var(--border);
                         font-variant-numeric: tabular-nums;
                         white-space: nowrap; }
#detail th { color: var(--text-3); font-weight: 600; text-align: right; }
#detail th:first-child, #detail td:first-child { text-align: left;
                                                 padding-left: 0; }
#detail tbody tr:hover td { background: var(--surface-alt); }
#detail .muted { color: var(--text-3); }

/* ---- right: report ---- */
#report { background: var(--surface-alt); }
#report pre { font-size: 12px; line-height: 1.5; margin: 0;
              padding: 12px 18px 60px; white-space: pre;
              color: var(--text-1); }
</style>
<div id="top">
  <div class="bar">
    <h1>TizenRT Heap Memory Map</h1>
    <label>Event <select id="eventSel"></select></label>
    <label>Heap <select id="heapSel"></select></label>
    <span class="legend">
      <span class="key"><span class="swatch heap"></span>heap</span>
      <span class="key"><span class="swatch stack"></span>stack</span>
      <span class="key"><span class="swatch free"></span>free</span>
      <span class="key"><span class="swatch unknown"></span>unknown</span>
    </span>
  </div>
  <div id="stats"></div>
</div>
<div id="main">
  <section class="panel">
    <div class="panel-title">Memory map &middot; top = low address</div>
    <div id="strip" class="panel-body"></div>
  </section>
  <div class="divider"></div>
  <section class="panel">
    <div class="panel-title">Selection detail</div>
    <div id="detail" class="panel-body"><span class="placeholder">click a
    block on the left to see who allocated it</span></div>
  </section>
  <div class="divider"></div>
  <section class="panel">
    <div class="panel-title">Report (report.txt)</div>
    <div id="report" class="panel-body"><pre id="reportPre"></pre></div>
  </section>
</div>
<script>
const data = __DATA__;
const strip = document.getElementById('strip');
const stats = document.getElementById('stats');
const detail = document.getElementById('detail');
const eventSel = document.getElementById('eventSel');
const heapSel = document.getElementById('heapSel');
let selectedRow = null;
document.getElementById('reportPre').textContent =
  data.report || '(no report)';

data.events.forEach((event, i) => {
  const opt = document.createElement('option');
  opt.value = i;
  opt.textContent = '#' + event.index + ' ' + event.kind +
    (event.command ? ' (' + event.command + ')' : '') +
    (event.timestamp ? ' @ ' + event.timestamp : '');
  eventSel.appendChild(opt);
});

function currentEvent() { return data.events[Number(eventSel.value)] || null; }
function currentHeap() {
  const event = currentEvent();
  return event ? event.heaps[Number(heapSel.value)] || null : null;
}
function rebuildHeapSel() {
  heapSel.innerHTML = '';
  const event = currentEvent();
  if (!event) return;
  event.heaps.forEach((heap, i) => {
    const opt = document.createElement('option');
    opt.value = i;
    opt.textContent = heap.name;
    heapSel.appendChild(opt);
  });
}
function fmt(size) {
  if (size >= 1048576) return (size / 1048576).toFixed(2) + ' MB';
  if (size >= 1024) return (size / 1024).toFixed(1) + ' KB';
  return size + ' B';
}
// Row height steps by histogram-style size bucket: bucket 0 = <=16 B,
// then one step per doubling (17-32, 33-64, ...). The minimum height
// always fits the one-line label, and each step is large enough that
// neighboring size classes are distinguishable at a glance.
function sizeBucket(size) {
  let bucket = 0, upper = 16;
  while (size > upper && bucket < 17) { upper *= 2; bucket += 1; }
  return bucket;   // 17 = anything above 1 MB
}
function rowHeight(run) {
  return 18 + sizeBucket(run.size) * 4;   // 18px .. 86px
}
function rowLabel(run) {
  const size = fmt(run.size);
  if (run.category === 'free') return run.address + '  FREE  ' + size;
  if (run.category === 'unknown')
    return run.address + '  UNKNOWN  ' + size;
  const count = run.count > 1 ? '  (' + run.count + ' allocs)' : '';
  return run.address + '  ' + run.category.toUpperCase() + '  ' + size +
         count;
}
function rowTitle(run) {
  return [
    'Start : ' + run.address,
    'Size  : ' + run.size + ' B (' + fmt(run.size) + ')',
    'Type  : ' + run.category,
    'Blocks: ' + run.count,
  ].join('\\n');
}
function esc(value) {
  return String(value == null ? '' : value)
    .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
}
// Raw addr2line result ("function location"), exactly as the tool
// printed it -- including "?? ??:0" for addresses it could not resolve.
// null means symbol resolution never ran (no binary): show owner only.
function codeOf(block) {
  if (block.function == null) return null;
  return block.function + ' ' + (block.location || '??:0');
}
function groupBy(blocks, keyOf, describe) {
  const groups = new Map();
  blocks.forEach(block => {
    const key = keyOf(block);
    const group = groups.get(key) ||
      Object.assign({count: 0, total: 0}, describe(block));
    group.count += 1;
    group.total += block.size;
    groups.set(key, group);
  });
  return [...groups.values()].sort((a, b) => b.total - a.total);
}
// rows: {label, sub?, subTitle?, count, total}; sub is a second, muted
// line in the first cell (used for the addr2line code location).
function aggTable(title, firstColumn, rows) {
  let html = '<h4>' + title + '</h4><table><thead><tr><th>' + firstColumn +
    '</th><th>Count</th><th>Total</th><th>%</th></tr></thead><tbody>';
  const total = rows.reduce((sum, row) => sum + row.total, 0);
  rows.forEach(row => {
    const percent = total ? (row.total * 100 / total).toFixed(1) : '0';
    let cell = esc(row.label);
    if (row.sub) {
      cell += '<br><span class="muted" title="' + esc(row.subTitle || '') +
        '">' + esc(row.sub) + '</span>';
    }
    html += '<tr><td>' + cell + '</td><td>' + row.count +
      '</td><td>' + esc(fmt(row.total)) + '</td><td>' + percent +
      '</td></tr>';
  });
  return html + '</tbody></table>';
}
function renderDetail(run) {
  const end = '0x' + (parseInt(run.address, 16) + run.size).toString(16);
  let html = '<h3>' + run.category.toUpperCase() + '  ' +
    esc(run.address) + ' .. ' + esc(end) + '</h3>' +
    '<div class="sub">size <b>' + esc(fmt(run.size)) + '</b> (' + run.size +
    ' B) &middot; ' + run.count + ' block(s)</div>';
  if (!run.blocks.length) {
    const note = {
      free: 'Free memory: nothing is allocated here.',
      unknown: 'Unknown area: not covered by parsed dump rows ' +
               '(truncated or corrupted dump).',
    };
    detail.innerHTML = html +
      '<span class="muted">' + (note[run.category] || '') + '</span>';
    return;
  }
  html += aggTable('By PID', 'PID (task)', groupBy(run.blocks,
    b => b.pid,
    b => ({label: (b.pid || '?') + (b.task ? ' (' + b.task + ')' : '')})));
  html += aggTable('By owner', 'Owner / Code', groupBy(run.blocks,
    b => b.owner,
    b => ({label: b.owner || '?', sub: codeOf(b)})));
  const hasCode = run.blocks.some(b => codeOf(b) !== null);
  html += '<h4>Blocks (address order)</h4><table><thead><tr>' +
    '<th>Address</th><th>Size</th><th>PID</th><th>Owner</th>' +
    (hasCode ? '<th style="text-align:left">Code</th>' : '') +
    '</tr></thead><tbody>';
  run.blocks.forEach(block => {
    html += '<tr>' +
      '<td>' + esc(block.address) + '</td><td>' + block.size + '</td>' +
      '<td>' + esc((block.pid || '?') +
                   (block.task ? ' (' + block.task + ')' : '')) + '</td>' +
      '<td>' + esc(block.owner || '') + '</td>' +
      (hasCode ? '<td style="text-align:left">' +
                 esc(codeOf(block) || '') + '</td>' : '') +
      '</tr>';
  });
  detail.innerHTML = html + '</tbody></table>';
}
function selectRun(row, run) {
  if (selectedRow) selectedRow.classList.remove('selected');
  selectedRow = row;
  row.classList.add('selected');
  renderDetail(run);
}
function renderStats(heap) {
  const free = heap.runs.filter(r => r.category === 'free');
  const unknown = heap.runs.filter(r => r.category === 'unknown');
  const freeTotal = free.reduce((sum, r) => sum + r.size, 0);
  const freeMax = free.reduce((max, r) => Math.max(max, r.size), 0);
  const parts = [];
  const s = heap.stats || {};
  if (s.fragmentation_index != null) {
    parts.push('fragmentation <b>' + s.fragmentation_index.toFixed(2) +
               '%</b>');
  }
  parts.push('free chunks in dump <b>' + free.length + '</b> (total ' +
             fmt(freeTotal) + ', largest ' + fmt(freeMax) + ')');
  if (s.free_nodes != null) {
    parts.push('reported free nodes <b>' + s.free_nodes + '</b>' +
               (s.largest_free != null
                ? ' (largest ' + fmt(s.largest_free) + ')' : ''));
  }
  if (unknown.length) {
    parts.push('unknown areas <b>' + unknown.length + '</b>');
  }
  let html = parts.map(p => '<span class="stat">' + p + '</span>').join('');
  if (heap.truncated) {
    html += '<span class="stat warn">(!) dump truncated: memory after ' +
            'the last parsed block is shown as unknown</span>';
  }
  stats.innerHTML = html;
}
function render() {
  strip.innerHTML = '';
  selectedRow = null;
  detail.innerHTML = '<span class="placeholder">click a block on the ' +
    'left to see who allocated it</span>';
  const heap = currentHeap();
  if (!heap) {
    stats.innerHTML = data.events.length
      ? '' : '<span class="warn">No heap dump (REGION table) in the log.</span>';
    return;
  }
  renderStats(heap);
  const regions = heap.regions.length ? heap.regions
    : [{number: heap.runs.length ? heap.runs[0].region : 0,
        start: '?', end: '?', size: 0}];
  regions.forEach(region => {
    const header = document.createElement('div');
    header.className = 'region-header';
    header.textContent = 'REGION #' + region.number + '  ' + region.start +
      ' .. ' + region.end + '  (' + fmt(region.size) + ')';
    strip.appendChild(header);
    heap.runs
      .filter(r => r.region === region.number)
      .forEach(run => {
        const row = document.createElement('div');
        row.className = 'run ' + run.category;
        row.style.height = rowHeight(run) + 'px';
        row.textContent = rowLabel(run);
        row.title = rowTitle(run);
        row.setAttribute('role', 'button');
        row.tabIndex = 0;
        row.addEventListener('click', () => selectRun(row, run));
        row.addEventListener('keydown', event => {
          if (event.key === 'Enter' || event.key === ' ') {
            event.preventDefault();
            selectRun(row, run);
          }
        });
        strip.appendChild(row);
      });
  });
}
eventSel.addEventListener('input', () => { rebuildHeapSel(); render(); });
heapSel.addEventListener('input', render);
rebuildHeapSel();
render();
</script>
"""


def report_html(result, output):
    """Write memory_map.html: vertical fragmentation map per event/heap."""
    events_payload = []
    for event in result.events:
        directory = _pid_directory(event)
        heaps_payload = []
        for heap in event.heaps:
            if not heap.regions and not heap.blocks:
                continue
            derived = heap.derived or DerivedMetrics()
            details = heap.details or HeapDetails()
            heaps_payload.append({
                "name": heap.name,
                "truncated": heap.blocks_truncated,
                "stats": {
                    "fragmentation_index": derived.fragmentation_index,
                    "free_nodes": details.free_node_count,
                    "largest_free": details.largest_free_node,
                },
                "regions": [asdict(region) for region in heap.regions],
                "runs": _html_runs(heap, directory),
            })
        if heaps_payload:
            events_payload.append({
                "index": event.index,
                "kind": event.kind,
                "command": event.command,
                "timestamp": event.time_start,
                "heaps": heaps_payload,
            })

    payload = {
        "source": result.source.path,
        "report": render_console_report(result),
        "events": events_payload,
    }
    data = json.dumps(payload, default=_json_default).replace("</", "<\\/")
    with open(output.path(HTML_MAP_NAME), "w", encoding="utf-8") as html:
        html.write(HTML_TEMPLATE.replace("__DATA__", data))


# ========================================================================
# [S8] CLI & main
# ========================================================================

def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Analyze TizenRT heap logs (alloc-fail dumps, heapinfo output)"
    )
    parser.add_argument("logfile", help="serial/console log file to analyze")
    parser.add_argument("--bin-dir", default="build/output/bin",
                        help="directory containing ELF binaries for addr2line")
    parser.add_argument("--addr2line", default="arm-none-eabi-addr2line",
                        help="addr2line command")
    parser.add_argument("--no-symbols", action="store_true",
                        help="skip addr2line symbol resolution")
    parser.add_argument("--no-excel", action="store_true",
                        help="skip report.xlsx generation")
    parser.add_argument("--no-html", action="store_true",
                        help="skip memory_map.html generation")
    parser.add_argument("--event", type=int, default=None, metavar="N",
                        help="report only the N-th event (default: all)")
    return parser


def main(argv=None):
    args = build_arg_parser().parse_args(argv)

    if not os.path.isfile(args.logfile):
        print(f"Error: log file '{args.logfile}' not found", file=sys.stderr)
        return 1

    try:
        output = OutputManager(DEFAULT_OUTPUT_DIR).prepare()
    except ValueError as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    extracted = extract_events(args.logfile, output)
    events = [parse_event(item) for item in extracted]
    for event in events:
        analyze_event(event)

    result = AnalysisResult(
        source=SourceInfo(
            path=os.path.abspath(args.logfile),
            size=os.path.getsize(args.logfile),
            total_events=len(events),
        ),
        events=events,
    )

    resolve_symbols(result, args.bin_dir, args.addr2line, args.no_symbols)

    if args.event is not None:
        if not 1 <= args.event <= len(result.events):
            print(
                f"Error: --event {args.event} out of range "
                f"(1..{len(result.events)})",
                file=sys.stderr,
            )
            return 1
        result.events = [result.events[args.event - 1]]

    report_console(result, output)
    report_json(result, output)
    if not args.no_excel:
        report_excel(result, output)
    if not args.no_html:
        report_html(result, output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
